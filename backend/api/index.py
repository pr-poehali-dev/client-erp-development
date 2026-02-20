import json
import os
import psycopg2
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
import calendar
import base64
import hashlib
import secrets
from io import BytesIO

def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'])

def safe_float(v, field_name='значение'):
    if v is None:
        raise ValueError('Не указано: %s' % field_name)
    return float(str(v).replace(',', '.'))

def safe_int(v, field_name='значение'):
    if v is None:
        raise ValueError('Не указано: %s' % field_name)
    return int(float(str(v).replace(',', '.')))

def last_day_of_month(d):
    return d.replace(day=calendar.monthrange(d.year, d.month)[1])

def add_months(d, months):
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)

def serialize(val):
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    return val

def query_rows(cur, sql):
    cur.execute(sql)
    cols = [d[0] for d in cur.description]
    return [{cols[i]: serialize(r[i]) for i in range(len(cols))} for r in cur.fetchall()]

def query_one(cur, sql):
    cur.execute(sql)
    if cur.rowcount == 0:
        return None
    cols = [d[0] for d in cur.description]
    row = cur.fetchone()
    return {cols[i]: serialize(row[i]) for i in range(len(cols))}

def calc_payment_dates(start_date, term):
    dates = []
    for i in range(1, term + 1):
        if i == term:
            dates.append(add_months(start_date, term))
        else:
            dates.append(last_day_of_month(add_months(start_date, i)))
    return dates

def calc_annuity_schedule(amount, rate, term, start_date):
    monthly_rate = Decimal(str(rate)) / Decimal('100') / Decimal('12')
    amt = Decimal(str(amount))
    if monthly_rate > 0:
        annuity = amt * monthly_rate * (1 + monthly_rate) ** term / ((1 + monthly_rate) ** term - 1)
    else:
        annuity = amt / term
    annuity = annuity.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    schedule = []
    balance = amt
    pay_dates = calc_payment_dates(start_date, term)
    for i in range(1, term + 1):
        payment_date = pay_dates[i - 1]
        interest = (balance * monthly_rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if i == term:
            principal = balance
            payment = principal + interest
        else:
            principal = annuity - interest
            if principal < 0:
                principal = Decimal('0')
            payment = annuity
        balance = balance - principal
        if balance < 0:
            balance = Decimal('0')
        schedule.append({
            'payment_no': i, 'payment_date': payment_date.isoformat(),
            'payment_amount': float(payment), 'principal_amount': float(principal),
            'interest_amount': float(interest), 'balance_after': float(balance),
        })
    return schedule, float(annuity)

def calc_end_of_term_schedule(amount, rate, term, start_date):
    amt = Decimal(str(amount))
    schedule = []
    balance = amt
    pay_dates = calc_payment_dates(start_date, term)
    for i in range(1, term + 1):
        payment_date = pay_dates[i - 1]
        prev_date = pay_dates[i - 2] if i > 1 else start_date
        days_in_period = (payment_date - prev_date).days
        interest = (balance * Decimal(str(rate)) / Decimal('100') * Decimal(str(days_in_period)) / Decimal('360')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if i == term:
            principal = balance
            payment = principal + interest
            balance = Decimal('0')
        else:
            principal = Decimal('0')
            payment = interest
        schedule.append({
            'payment_no': i, 'payment_date': payment_date.isoformat(),
            'payment_amount': float(payment), 'principal_amount': float(principal),
            'interest_amount': float(interest), 'balance_after': float(balance),
        })
    return schedule, float(schedule[0]['payment_amount']) if schedule else 0

def calc_savings_schedule(amount, rate, term, start_date, payout_type):
    amt = Decimal(str(amount))
    schedule = []
    cumulative = Decimal('0')
    close_date = add_months(start_date, term)
    for i in range(1, term + 1):
        if i == 1:
            period_start = start_date
        else:
            period_start = last_day_of_month(add_months(start_date, i - 1))
        if i == term:
            period_end = close_date
        else:
            period_end = last_day_of_month(add_months(start_date, i))
        actual_days = (period_end - period_start).days
        interest = (amt * Decimal(str(rate)) / Decimal('100') * Decimal(str(actual_days)) / Decimal('365')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        cumulative += interest
        balance_after = float(amt + cumulative) if payout_type == 'end_of_term' else float(amt)
        schedule.append({
            'period_no': i, 'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat(), 'interest_amount': float(interest),
            'cumulative_interest': float(cumulative), 'balance_after': balance_after,
        })
    return schedule

def refresh_loan_overdue_status(cur, lid):
    cur.execute("SELECT status FROM loans WHERE id=%s" % lid)
    row = cur.fetchone()
    if not row or row[0] == 'closed':
        return
    
    cur.execute("""
        SELECT COUNT(*) FROM loan_schedule
        WHERE loan_id=%s AND status IN ('pending', 'overdue')
          AND payment_date < CURRENT_DATE
    """ % lid)
    has_overdue = cur.fetchone()[0] > 0
    
    if has_overdue:
        cur.execute("UPDATE loans SET status='overdue', updated_at=NOW() WHERE id=%s AND status='active'" % lid)
        cur.execute("""
            UPDATE loan_schedule SET status='overdue', overdue_days=(CURRENT_DATE - payment_date)
            WHERE loan_id=%s AND status IN ('pending') AND payment_date < CURRENT_DATE
        """ % lid)
    else:
        cur.execute("UPDATE loans SET status='active', updated_at=NOW() WHERE id=%s AND status='overdue'" % lid)
        cur.execute("UPDATE loan_schedule SET overdue_days=0 WHERE loan_id=%s AND status='overdue'" % lid)
        cur.execute("UPDATE loan_schedule SET status='pending' WHERE loan_id=%s AND status='overdue'" % lid)
    
    cur.execute("""
        UPDATE loan_schedule SET overdue_days=(CURRENT_DATE - payment_date)
        WHERE loan_id=%s AND status = 'partial' AND payment_date < CURRENT_DATE
    """ % lid)

def recalc_loan_schedule_statuses(cur, lid):
    cur.execute("UPDATE loan_schedule SET paid_amount=0, paid_date=NULL, status='pending' WHERE loan_id=%s" % lid)
    cur.execute("SELECT id, payment_date, amount FROM loan_payments WHERE loan_id=%s ORDER BY payment_date, id" % lid)
    payments = cur.fetchall()
    for pay in payments:
        pay_id = pay[0]
        pay_date = str(pay[1])
        remaining = Decimal(str(pay[2]))
        pay_pp = Decimal('0')
        pay_ip = Decimal('0')
        pay_pnp = Decimal('0')
        
        cur.execute("""
            SELECT id, principal_amount, interest_amount, penalty_amount, paid_amount
            FROM loan_schedule WHERE loan_id=%s AND status IN ('pending','partial')
            ORDER BY payment_no, id
        """ % lid)
        unpaid_rows = cur.fetchall()
        
        for row in unpaid_rows:
            if remaining <= Decimal('0.005'):
                break
            sid = row[0]
            sp = Decimal(str(row[1]))
            si = Decimal(str(row[2]))
            spn = Decimal(str(row[3]))
            spa = Decimal(str(row[4]))
            
            already_i = min(spa, si)
            already_pn = min(spa - si, spn) if spa > si else Decimal('0')
            already_pp = spa - already_i - already_pn if spa > already_i + already_pn else Decimal('0')
            
            need_i = si - already_i
            need_pn = spn - already_pn
            need_pp = sp - already_pp
            need_total = need_i + need_pn + need_pp

            if need_total <= Decimal('0.005'):
                continue

            # Берём ровно столько сколько нужно для закрытия периода.
            # Если денег меньше чем нужно — частичное покрытие.
            # Если денег больше — берём только need_total, остаток пойдёт в ОД.
            take_total = min(remaining, need_total)
            item_i = min(take_total, need_i)
            after_i = take_total - item_i
            item_pn = min(after_i, need_pn)
            item_pp = after_i - item_pn
            remaining -= take_total
            
            pay_ip += item_i
            pay_pnp += item_pn
            pay_pp += item_pp
            
            total_item = sp + si + spn
            new_paid = spa + item_i + item_pn + item_pp
            ns = 'paid' if new_paid >= total_item else 'partial'
            cur.execute("UPDATE loan_schedule SET paid_amount=%s, paid_date='%s', status='%s' WHERE id=%s" % (float(new_paid), pay_date, ns, sid))

            # Остаток платежа переходит на следующие периоды в цикле
        
        cur.execute("UPDATE loan_payments SET principal_part=%s, interest_part=%s, penalty_part=%s WHERE id=%s" % (
            float(pay_pp), float(pay_ip), float(pay_pnp), pay_id))
    
    refresh_loan_overdue_status(cur, lid)

def esc(val):
    return str(val).replace("'", "''") if val else ''

def audit_log(cur, staff, action, entity, entity_id=None, entity_label='', details='', ip=''):
    uid = staff.get('user_id') if staff else None
    uname = esc(staff.get('name', '')) if staff else ''
    urole = staff.get('role', '') if staff else ''
    cur.execute("INSERT INTO audit_log (user_id, user_name, user_role, action, entity, entity_id, entity_label, details, ip) VALUES (%s, '%s', '%s', '%s', '%s', %s, '%s', '%s', '%s')" % (
        uid or 'NULL', uname, urole, esc(action), esc(entity),
        entity_id or 'NULL', esc(entity_label), esc(details), esc(ip)
    ))

def handle_members(method, params, body, cur, conn, staff=None, ip=''):
    if method == 'GET':
        member_id = params.get('id')
        if member_id:
            return query_one(cur, "SELECT * FROM members WHERE id = %s" % member_id)
        return query_rows(cur, """
            SELECT m.id, m.member_no, m.member_type,
                   CASE WHEN m.member_type = 'FL' THEN CONCAT(m.last_name, ' ', m.first_name, ' ', m.middle_name)
                        ELSE m.company_name END as name,
                   m.inn, m.phone, m.email, m.status, m.created_at,
                   (SELECT COUNT(*) FROM loans l WHERE l.member_id = m.id AND l.status != 'closed') as active_loans,
                   (SELECT COUNT(*) FROM savings s WHERE s.member_id = m.id AND s.status = 'active') as active_savings
            FROM members m WHERE m.status != 'deleted' ORDER BY m.created_at DESC
        """)

    elif method == 'POST':
        mt = body.get('member_type', 'FL')
        cur.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM members")
        next_id = cur.fetchone()[0]
        member_no = 'П-%06d' % next_id

        if mt == 'FL':
            cur.execute("""
                INSERT INTO members (member_no, member_type, last_name, first_name, middle_name,
                    birth_date, birth_place, inn, passport_series, passport_number,
                    passport_dept_code, passport_issue_date, passport_issued_by,
                    registration_address, phone, email, telegram, bank_bik, bank_account,
                    marital_status, spouse_fio, spouse_phone, extra_phone, extra_contact_fio)
                VALUES ('%s', 'FL', '%s', '%s', '%s', %s, '%s', '%s', '%s', '%s', '%s', %s, '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')
                RETURNING id, member_no
            """ % (
                member_no, esc(body.get('last_name')), esc(body.get('first_name')), esc(body.get('middle_name')),
                ("'%s'" % body['birth_date']) if body.get('birth_date') else 'NULL',
                esc(body.get('birth_place')), esc(body.get('inn')),
                esc(body.get('passport_series')), esc(body.get('passport_number')),
                esc(body.get('passport_dept_code')),
                ("'%s'" % body['passport_issue_date']) if body.get('passport_issue_date') else 'NULL',
                esc(body.get('passport_issued_by')), esc(body.get('registration_address')),
                esc(body.get('phone')), esc(body.get('email')), esc(body.get('telegram')),
                esc(body.get('bank_bik')), esc(body.get('bank_account')),
                esc(body.get('marital_status')), esc(body.get('spouse_fio')),
                esc(body.get('spouse_phone')), esc(body.get('extra_phone')), esc(body.get('extra_contact_fio')),
            ))
        else:
            cur.execute("""
                INSERT INTO members (member_no, member_type, inn, company_name, director_fio,
                    director_phone, contact_person_fio, contact_person_phone, bank_bik, bank_account)
                VALUES ('%s', 'UL', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')
                RETURNING id, member_no
            """ % (
                member_no, esc(body.get('inn')), esc(body.get('company_name')),
                esc(body.get('director_fio')), esc(body.get('director_phone')),
                esc(body.get('contact_person_fio')), esc(body.get('contact_person_phone')),
                esc(body.get('bank_bik')), esc(body.get('bank_account')),
            ))
        result = cur.fetchone()
        label = body.get('last_name', body.get('company_name', ''))
        audit_log(cur, staff, 'create', 'member', result[0], '%s %s' % (result[1], label), '', ip)
        conn.commit()
        return {'id': result[0], 'member_no': result[1]}

    elif method == 'PUT':
        member_id = body.get('id')
        updates = []
        for f in ['last_name','first_name','middle_name','birth_place','inn','passport_series','passport_number',
                   'passport_dept_code','passport_issued_by','registration_address','phone','email','telegram',
                   'bank_bik','bank_account','marital_status','spouse_fio','spouse_phone','extra_phone',
                   'extra_contact_fio','company_name','director_fio','director_phone','contact_person_fio',
                   'contact_person_phone','status']:
            if f in body:
                updates.append("%s = '%s'" % (f, esc(body[f])))
        for f in ['birth_date','passport_issue_date']:
            if f in body and body[f]:
                updates.append("%s = '%s'" % (f, body[f]))
        if updates:
            updates.append("updated_at = NOW()")
            cur.execute("UPDATE members SET %s WHERE id = %s" % (', '.join(updates), member_id))
            changed = [f for f in body if f not in ('id', 'entity')]
            audit_log(cur, staff, 'update', 'member', member_id, '', ', '.join(changed), ip)
            conn.commit()
        return {'success': True}

    elif method == 'DELETE':
        member_id = params.get('id')
        cur.execute("SELECT member_no, CASE WHEN member_type='FL' THEN CONCAT(last_name,' ',first_name,' ',middle_name) ELSE company_name END as name FROM members WHERE id=%s" % member_id)
        mem = cur.fetchone()
        if not mem:
            return {'error': 'Пайщик не найден'}
        member_no, member_name = mem[0], mem[1]
        cur.execute("SELECT COUNT(*) FROM loans WHERE member_id=%s AND status='active'" % member_id)
        if cur.fetchone()[0] > 0:
            return {'error': 'Нельзя удалить пайщика с активными займами'}
        cur.execute("SELECT COUNT(*) FROM savings WHERE member_id=%s AND status='active'" % member_id)
        if cur.fetchone()[0] > 0:
            return {'error': 'Нельзя удалить пайщика с активными сбережениями'}
        cur.execute("SELECT COUNT(*) FROM share_accounts WHERE member_id=%s AND status='active'" % member_id)
        if cur.fetchone()[0] > 0:
            return {'error': 'Нельзя удалить пайщика с активными паевыми счетами'}
        cur.execute("DELETE FROM members WHERE id=%s" % member_id)
        audit_log(cur, staff, 'delete', 'member', member_id, member_no, member_name, ip)
        conn.commit()
        return {'success': True}

def handle_loans(method, params, body, cur, conn, staff=None, ip=''):
    if method == 'GET':
        action = params.get('action', 'list')
        if action == 'detail':
            loan = query_one(cur, "SELECT * FROM loans WHERE id = %s" % params['id'])
            if not loan:
                return None
            cur.execute("SELECT CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name) ELSE m.company_name END FROM members m WHERE m.id=%s" % loan['member_id'])
            nr = cur.fetchone()
            loan['member_name'] = nr[0] if nr else ''
            if loan.get('org_id'):
                org_row = query_one(cur, "SELECT name, short_name FROM organizations WHERE id=%s" % loan['org_id'])
                loan['org_name'] = org_row['name'] if org_row else ''
                loan['org_short_name'] = org_row['short_name'] if org_row else ''
            loan['schedule'] = query_rows(cur, "SELECT * FROM loan_schedule WHERE loan_id=%s ORDER BY payment_no" % params['id'])
            loan['payments'] = query_rows(cur, "SELECT * FROM loan_payments WHERE loan_id=%s ORDER BY payment_date" % params['id'])
            return loan
        elif action == 'check_status':
            loan_number = params.get('loan_number')
            if not loan_number:
                return {'error': 'Не указан номер договора'}
            cur.execute("SELECT id FROM loans WHERE contract_no = '%s'" % esc(loan_number))
            lr = cur.fetchone()
            if not lr:
                return {'error': 'Договор не найден'}
            lid = lr[0]
            
            cur.execute("SELECT payment_no, payment_date, principal_amount, interest_amount, penalty_amount, COALESCE(paid_amount,0) as paid_amount, status, paid_date FROM loan_schedule WHERE loan_id=%s ORDER BY payment_no" % lid)
            schedule = [{'payment_no': r[0], 'payment_date': str(r[1]), 'principal': float(r[2]), 'interest': float(r[3]), 'penalty': float(r[4]), 'paid_amount': float(r[5]), 'status': r[6], 'paid_date': str(r[7]) if r[7] else None} for r in cur.fetchall()]
            
            cur.execute("SELECT payment_date, amount, principal_part, interest_part, penalty_part FROM loan_payments WHERE loan_id=%s ORDER BY payment_date" % lid)
            payments = [{'payment_date': str(r[0]), 'amount': float(r[1]), 'principal': float(r[2]), 'interest': float(r[3]), 'penalty': float(r[4])} for r in cur.fetchall()]
            
            total_paid_from_schedule = sum(s['paid_amount'] for s in schedule)
            total_paid_from_payments = sum(p['amount'] for p in payments)
            
            last_paid = [s for s in schedule if s['status'] == 'paid']
            last_paid_period = last_paid[-1]['payment_date'] if last_paid else None
            
            return {
                'loan_id': lid,
                'loan_number': loan_number,
                'schedule': schedule,
                'payments': payments,
                'total_paid_from_schedule': total_paid_from_schedule,
                'total_paid_from_payments': total_paid_from_payments,
                'last_paid_period': last_paid_period,
                'stats': {
                    'paid': len([s for s in schedule if s['status'] == 'paid']),
                    'partial': len([s for s in schedule if s['status'] == 'partial']),
                    'pending': len([s for s in schedule if s['status'] == 'pending']),
                    'overdue': len([s for s in schedule if s['status'] == 'overdue'])
                }
            }
        elif action == 'reconciliation_report':
            lid = params.get('id')
            if not lid:
                return {'error': 'Не указан id договора'}
            loan = query_one(cur, "SELECT * FROM loans WHERE id=%s" % lid)
            if not loan:
                return {'error': 'Договор не найден'}
            cur.execute("SELECT CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name) ELSE m.company_name END FROM members m WHERE m.id=%s" % loan['member_id'])
            nr = cur.fetchone()
            loan['member_name'] = nr[0] if nr else ''

            cur.execute("""
                SELECT id, payment_no, payment_date, payment_amount, principal_amount, interest_amount,
                       COALESCE(penalty_amount,0) as penalty_amount, COALESCE(paid_amount,0) as paid_amount,
                       status, paid_date
                FROM loan_schedule WHERE loan_id=%s ORDER BY payment_no, id
            """ % lid)
            schedule_rows = cur.fetchall()

            cur.execute("""
                SELECT id, payment_date, amount, principal_part, interest_part, COALESCE(penalty_part,0) as penalty_part, payment_type
                FROM loan_payments WHERE loan_id=%s ORDER BY payment_date, id
            """ % lid)
            payment_rows = cur.fetchall()

            schedule_list = []
            for r in schedule_rows:
                sch_id, pno, pdate, pamt, princ, inter, penal, paid, status, paid_date = r
                total = float(Decimal(str(princ)) + Decimal(str(inter)) + Decimal(str(penal)))
                schedule_list.append({
                    'id': sch_id,
                    'payment_no': pno,
                    'plan_date': str(pdate),
                    'plan_amount': float(pamt),
                    'plan_principal': float(princ),
                    'plan_interest': float(inter),
                    'plan_penalty': float(penal),
                    'plan_total': round(total, 2),
                    'paid_amount': float(paid),
                    'status': status,
                    'paid_date': str(paid_date) if paid_date else None,
                    'payments': []
                })

            sch_by_id = {s['id']: s for s in schedule_list}

            # Привязываем каждый платёж к периоду(ам) графика через paid_amount на графике.
            # Используем реальные данные principal_part/interest_part из loan_payments —
            # они были рассчитаны в момент внесения на актуальном графике.
            # Для сопоставления: идём хронологически, каждый платёж закрывает
            # ближайший незакрытый период(ы) в порядке payment_no.
            schedule_remaining = {}  # schedule_id -> сколько ещё не покрыто платежами
            for sch in schedule_list:
                total_plan = Decimal(str(sch['plan_principal'])) + Decimal(str(sch['plan_interest'])) + Decimal(str(sch['plan_penalty']))
                schedule_remaining[sch['id']] = total_plan

            for pay_row in payment_rows:
                pay_id, pay_date, pay_amt, pay_pp, pay_ip, pay_pnp, pay_type = pay_row
                remaining = Decimal(str(pay_amt))
                # Распределяем по периодам в порядке payment_no, используя реальную разбивку ОД/% из платежа
                pay_pp_d = Decimal(str(pay_pp))
                pay_ip_d = Decimal(str(pay_ip))
                pay_pnp_d = Decimal(str(pay_pnp))
                # Сколько ОД/% осталось распределить по периодам
                pp_left = pay_pp_d
                ip_left = pay_ip_d
                pnp_left = pay_pnp_d

                for sch in schedule_list:
                    if remaining <= Decimal('0.005'):
                        break
                    s_id = sch['id']
                    need = schedule_remaining.get(s_id, Decimal('0'))
                    if need <= Decimal('0.005'):
                        continue

                    take = min(remaining, need)
                    remaining -= take
                    schedule_remaining[s_id] = need - take

                    # Пропорционально распределяем ОД/% в рамках взятой суммы
                    ratio = take / Decimal(str(pay_amt)) if Decimal(str(pay_amt)) > 0 else Decimal('1')
                    take_pp = min(pp_left, (pay_pp_d * ratio).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    take_ip = min(ip_left, (pay_ip_d * ratio).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    take_pnp = min(pnp_left, (pay_pnp_d * ratio).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    # Последняя часть берёт остаток чтобы сумма сошлась
                    if remaining <= Decimal('0.005'):
                        take_pp = pp_left
                        take_ip = ip_left
                        take_pnp = pnp_left
                    pp_left -= take_pp
                    ip_left -= take_ip
                    pnp_left -= take_pnp

                    sch_by_id[s_id]['payments'].append({
                        'payment_id': pay_id,
                        'fact_date': str(pay_date),
                        'amount': round(float(take), 2),
                        'principal': round(float(take_pp), 2),
                        'interest': round(float(take_ip), 2),
                        'penalty': round(float(take_pnp), 2),
                        'payment_type': pay_type,
                    })

            total_plan = sum(s['plan_total'] for s in schedule_list)
            total_paid = sum(float(r[2]) for r in payment_rows)
            total_overdue = sum(s['plan_total'] - s['paid_amount'] for s in schedule_list if s['status'] in ('overdue', 'partial') and s['plan_date'] < date.today().isoformat())

            return {
                'loan': {
                    'id': loan['id'],
                    'contract_no': loan['contract_no'],
                    'member_name': loan['member_name'],
                    'amount': float(loan['amount']),
                    'rate': float(loan['rate']),
                    'term_months': loan['term_months'],
                    'start_date': str(loan['start_date']),
                    'end_date': str(loan['end_date']),
                    'status': loan['status'],
                    'balance': float(loan['balance']),
                },
                'schedule': schedule_list,
                'summary': {
                    'total_plan': round(total_plan, 2),
                    'total_paid': round(total_paid, 2),
                    'total_diff': round(total_plan - total_paid, 2),
                    'total_overdue': round(total_overdue, 2),
                    'periods_total': len(schedule_list),
                    'periods_paid': len([s for s in schedule_list if s['status'] == 'paid']),
                    'periods_partial': len([s for s in schedule_list if s['status'] == 'partial']),
                    'periods_overdue': len([s for s in schedule_list if s['status'] == 'overdue']),
                    'periods_pending': len([s for s in schedule_list if s['status'] == 'pending']),
                }
            }
        elif action == 'schedule':
            a, r, t = safe_float(params['amount'], 'сумма'), safe_float(params['rate'], 'ставка'), safe_int(params['term'], 'срок')
            st = params.get('schedule_type', 'annuity')
            sd = date.fromisoformat(params.get('start_date', date.today().isoformat()))
            fn = calc_annuity_schedule if st == 'annuity' else calc_end_of_term_schedule
            schedule, monthly = fn(a, r, t, sd)
            return {'schedule': schedule, 'monthly_payment': monthly}
        else:
            return query_rows(cur, """
                SELECT l.id, l.contract_no, l.amount, l.rate, l.term_months, l.schedule_type,
                       l.start_date, l.end_date, l.monthly_payment, l.balance, l.status,
                       l.org_id,
                       CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name)
                            ELSE m.company_name END as member_name, m.id as member_id,
                       o.name as org_name, o.short_name as org_short_name
                FROM loans l JOIN members m ON m.id=l.member_id
                LEFT JOIN organizations o ON o.id=l.org_id
                ORDER BY l.created_at DESC
            """)

    elif method == 'POST':
        action = body.get('action', 'create')
        if action == 'create':
            cn = body['contract_no']
            mid = int(body['member_id'])
            a, r, t = safe_float(body['amount'], 'сумма'), safe_float(body['rate'], 'ставка'), safe_int(body['term_months'], 'срок')
            st = body.get('schedule_type', 'annuity')
            sd = date.fromisoformat(body.get('start_date', date.today().isoformat()))
            ed = add_months(sd, t)
            fn = calc_annuity_schedule if st == 'annuity' else calc_end_of_term_schedule
            schedule, monthly = fn(a, r, t, sd)
            org_id = body.get('org_id')

            cur.execute("""
                INSERT INTO loans (contract_no, member_id, amount, rate, term_months, schedule_type,
                    start_date, end_date, monthly_payment, balance, status, org_id)
                VALUES ('%s', %s, %s, %s, %s, '%s', '%s', '%s', %s, %s, 'active', %s) RETURNING id
            """ % (esc(cn), mid, a, r, t, st, sd.isoformat(), ed.isoformat(), monthly, a, org_id if org_id else 'NULL'))
            lid = cur.fetchone()[0]
            for item in schedule:
                cur.execute("""
                    INSERT INTO loan_schedule (loan_id, payment_no, payment_date, payment_amount,
                        principal_amount, interest_amount, balance_after)
                    VALUES (%s, %s, '%s', %s, %s, %s, %s)
                """ % (lid, item['payment_no'], item['payment_date'], item['payment_amount'],
                       item['principal_amount'], item['interest_amount'], item['balance_after']))
            audit_log(cur, staff, 'create', 'loan', lid, cn, 'Сумма: %s, ставка: %s%%, срок: %s мес.' % (a, r, t), ip)
            conn.commit()
            return {'id': lid, 'schedule': schedule, 'monthly_payment': monthly}

        elif action == 'payment':
            lid = int(body['loan_id'])
            pd = body['payment_date']
            amt = Decimal(str(safe_float(body['amount'], 'сумма')))
            overpay_strategy = body.get('overpay_strategy', '')

            cur.execute("SELECT balance, rate, schedule_type, term_months, monthly_payment, start_date FROM loans WHERE id = %s" % lid)
            loan_row = cur.fetchone()
            loan_bal = Decimal(str(loan_row[0]))
            l_rate, l_stype = float(loan_row[1]), loan_row[2]
            old_monthly = Decimal(str(loan_row[4]))
            loan_start = loan_row[5]

            cur.execute("""
                SELECT id, principal_amount, interest_amount, penalty_amount, paid_amount, payment_date, payment_no
                FROM loan_schedule WHERE loan_id=%s AND status IN ('pending','partial','overdue')
                ORDER BY payment_no, id LIMIT 1
            """ % lid)
            first_row = cur.fetchone()
            
            cur.execute("""
                SELECT COALESCE(SUM(principal_amount + interest_amount + penalty_amount - COALESCE(paid_amount, 0)), 0)
                FROM loan_schedule WHERE loan_id=%s AND status IN ('overdue', 'partial')
            """ % lid)
            total_overdue = Decimal(str(cur.fetchone()[0]))
            
            if first_row:
                f_sid = first_row[0]
                f_sp = Decimal(str(first_row[1]))
                f_si = Decimal(str(first_row[2]))
                f_spn = Decimal(str(first_row[3]))
                f_spa = Decimal(str(first_row[4]))
                f_pay_date = first_row[5]
                f_pay_no = first_row[6]
                first_payment_owed = f_sp + f_si + f_spn - f_spa
                current_owed = max(total_overdue, first_payment_owed)
            else:
                current_owed = loan_bal
                f_pay_date = date.fromisoformat(pd)
                f_pay_no = 0

            overpay_amount = amt - current_owed if amt > current_owed else Decimal('0')
            overpay_threshold = old_monthly * Decimal('0.5')
            is_significant_overpay = overpay_amount > overpay_threshold and amt < loan_bal and total_overdue == 0

            if is_significant_overpay and not overpay_strategy:
                need_i = f_si - min(f_spa, f_si)
                need_pn = f_spn - (min(f_spa - f_si, f_spn) if f_spa > f_si else Decimal('0'))
                total_principal = amt - need_i - need_pn
                if total_principal < 0:
                    total_principal = Decimal('0')
                new_bal_est = loan_bal - total_principal
                if new_bal_est < 0:
                    new_bal_est = Decimal('0')
                cur.execute("SELECT COUNT(*) FROM loan_schedule WHERE loan_id=%s AND status IN ('pending','partial','overdue')" % lid)
                remaining_periods = cur.fetchone()[0]
                remaining_after = remaining_periods - 1
                f_pay_date_parsed = date.fromisoformat(str(f_pay_date)) if isinstance(f_pay_date, str) else f_pay_date
                options = {}
                if remaining_after >= 1 and float(new_bal_est) > 0:
                    fn = calc_annuity_schedule if l_stype == 'annuity' else calc_end_of_term_schedule
                    sched_rp, monthly_rp = fn(float(new_bal_est), l_rate, remaining_after, f_pay_date_parsed)
                    options['reduce_payment'] = {'new_monthly': monthly_rp, 'new_term': remaining_after, 'description': 'Уменьшить ежемесячный платёж, срок останется прежним'}
                    best_term = remaining_after
                    for t in range(1, remaining_after + 1):
                        _, m = fn(float(new_bal_est), l_rate, t, f_pay_date_parsed)
                        if m <= float(old_monthly) * 1.1:
                            best_term = t
                            break
                    if best_term >= remaining_after:
                        best_term = max(remaining_after - 1, 1)
                    sched_rt, monthly_rt = fn(float(new_bal_est), l_rate, max(best_term, 1), f_pay_date_parsed)
                    options['reduce_term'] = {'new_monthly': monthly_rt, 'new_term': best_term, 'description': 'Сократить срок, платёж останется примерно прежним'}
                return {
                    'needs_choice': True,
                    'overpay_amount': float(overpay_amount),
                    'current_payment': float(current_owed),
                    'total_amount': float(amt),
                    'options': options,
                }

            pp = i_p = pnp = Decimal('0')
            remaining_amt = amt

            if first_row:
                cur.execute("""
                    SELECT id, principal_amount, interest_amount, penalty_amount, paid_amount
                    FROM loan_schedule WHERE loan_id=%s AND status IN ('pending','partial','overdue')
                    ORDER BY payment_no, id
                """ % lid)
                unpaid_rows = cur.fetchall()
                
                for row in unpaid_rows:
                    if remaining_amt <= Decimal('0.005'):
                        break
                    sid = row[0]
                    sp = Decimal(str(row[1]))
                    si = Decimal(str(row[2]))
                    spn = Decimal(str(row[3]))
                    spa = Decimal(str(row[4]))
                    
                    already_i = min(spa, si)
                    already_pn = min(spa - si, spn) if spa > si else Decimal('0')
                    already_pp = spa - already_i - already_pn if spa > already_i + already_pn else Decimal('0')
                    
                    need_i = si - already_i
                    need_pn = spn - already_pn
                    need_pp = sp - already_pp
                    need_total = need_i + need_pn + need_pp

                    if need_total <= Decimal('0.005'):
                        continue

                    take_total = min(remaining_amt, need_total)
                    item_i = min(take_total, need_i)
                    after_i = take_total - item_i
                    item_pn = min(after_i, need_pn)
                    item_pp = after_i - item_pn
                    remaining_amt -= take_total
                    
                    i_p += item_i
                    pnp += item_pn
                    pp += item_pp
                    
                    total_item = sp + si + spn
                    new_paid = spa + item_i + item_pn + item_pp
                    ns = 'paid' if new_paid >= total_item else 'partial'
                    cur.execute("UPDATE loan_schedule SET paid_amount=%s, paid_date='%s', status='%s' WHERE id=%s" % (float(new_paid), pd, ns, sid))

                    # Период закрыт и есть остаток — переплата идёт в ОД, не в % следующего периода
                    if ns == 'paid' and remaining_amt > Decimal('0.005'):
                        pp += remaining_amt
                        remaining_amt = Decimal('0')
                        break
            else:
                pp = min(amt, loan_bal)

            cur.execute("""
                INSERT INTO loan_payments (loan_id, payment_date, amount, principal_part, interest_part, penalty_part, payment_type)
                VALUES (%s, '%s', %s, %s, %s, %s, 'regular')
            """ % (lid, pd, float(amt), float(pp), float(i_p), float(pnp)))

            nb = loan_bal - pp
            if nb < 0: nb = Decimal('0')
            cur.execute("UPDATE loans SET balance=%s, updated_at=NOW() WHERE id=%s" % (float(nb), lid))
            if nb == 0:
                cur.execute("UPDATE loans SET status='closed', updated_at=NOW() WHERE id=%s" % lid)

            recalc_schedule = None
            auto_recalculated = False

            # Пересчёт нужен если была любая переплата (pp > плановый ОД периода)
            actual_overpay = pp - (f_sp - f_spa if first_row else Decimal('0')) if first_row else Decimal('0')
            if actual_overpay < 0:
                actual_overpay = Decimal('0')
            should_recalc = nb > 0 and (overpay_amount > 0 or actual_overpay > Decimal('0.005'))

            if should_recalc:
                cur.execute("SELECT COUNT(*) FROM loan_schedule WHERE loan_id=%s AND status IN ('pending','partial','overdue')" % lid)
                remaining_periods = cur.fetchone()[0]
                if remaining_periods > 0:
                    # Берём дату последнего оплаченного периода как базу для пересчёта
                    cur.execute("""
                        SELECT payment_date FROM loan_schedule
                        WHERE loan_id=%s AND status='paid'
                        ORDER BY payment_no DESC LIMIT 1
                    """ % lid)
                    last_paid_row = cur.fetchone()
                    if last_paid_row:
                        last_paid_date = last_paid_row[0] if isinstance(last_paid_row[0], date) else date.fromisoformat(str(last_paid_row[0]))
                    else:
                        last_paid_date = date.fromisoformat(pd)
                    cur.execute("DELETE FROM loan_schedule WHERE loan_id=%s AND status IN ('pending','partial','overdue')" % lid)
                    fn = calc_annuity_schedule if l_stype == 'annuity' else calc_end_of_term_schedule

                    if is_significant_overpay and overpay_strategy == 'reduce_term':
                        best_term = remaining_periods
                        for t in range(1, remaining_periods + 1):
                            _, m = fn(float(nb), l_rate, t, last_paid_date)
                            if m <= float(old_monthly) * 1.1:
                                best_term = t
                                break
                        if best_term >= remaining_periods:
                            best_term = max(remaining_periods - 1, 1)
                        new_sched, new_monthly = fn(float(nb), l_rate, max(best_term, 1), last_paid_date)
                    else:
                        new_sched, new_monthly = fn(float(nb), l_rate, remaining_periods, last_paid_date)
                        if not is_significant_overpay:
                            auto_recalculated = True

                    cur.execute("SELECT MAX(payment_no) FROM loan_schedule WHERE loan_id=%s" % lid)
                    max_no_row = cur.fetchone()
                    max_no = max_no_row[0] if max_no_row and max_no_row[0] else 0
                    for item in new_sched:
                        cur.execute("INSERT INTO loan_schedule (loan_id,payment_no,payment_date,payment_amount,principal_amount,interest_amount,balance_after) VALUES (%s,%s,'%s',%s,%s,%s,%s)" % (lid, max_no + item['payment_no'], item['payment_date'], item['payment_amount'], item['principal_amount'], item['interest_amount'], item['balance_after']))
                    ne = date.fromisoformat(new_sched[-1]['payment_date'])
                    new_term = max_no + len(new_sched)
                    cur.execute("UPDATE loans SET monthly_payment=%s, end_date='%s', term_months=%s, updated_at=NOW() WHERE id=%s" % (new_monthly, ne.isoformat(), new_term, lid))
                    recalc_schedule = new_sched

            if nb > 0:
                refresh_loan_overdue_status(cur, lid)

            pay_detail = 'Сумма: %s, ОД: %s, %%: %s' % (float(amt), float(pp), float(i_p))
            if recalc_schedule:
                pay_detail += ', график пересчитан (%s)' % (overpay_strategy if overpay_strategy else 'авто')
            audit_log(cur, staff, 'payment', 'loan', lid, '', pay_detail, ip)
            conn.commit()
            result = {'success': True, 'new_balance': float(nb), 'principal_part': float(pp), 'interest_part': float(i_p), 'penalty_part': float(pnp)}
            if recalc_schedule:
                result['schedule_recalculated'] = True
                result['new_monthly'] = new_monthly
            if auto_recalculated:
                result['auto_recalculated'] = True
            return result

        elif action == 'early_repayment':
            lid = int(body['loan_id'])
            amt = safe_float(body['amount'], 'сумма')
            rt = body.get('repayment_type', 'reduce_term')
            pd = body.get('payment_date', date.today().isoformat())

            cur.execute("SELECT amount, rate, balance, term_months, start_date, schedule_type, monthly_payment FROM loans WHERE id=%s" % lid)
            lr = cur.fetchone()
            cb, r, st = float(lr[2]), float(lr[1]), lr[5]
            old_monthly = float(lr[6]) if lr[6] else 0
            nb = cb - amt

            if nb <= 0:
                cur.execute("UPDATE loans SET balance=0, status='closed', updated_at=NOW() WHERE id=%s" % lid)
                cur.execute("UPDATE loan_schedule SET status='paid' WHERE loan_id=%s AND status IN ('pending','partial','overdue')" % lid)
                cur.execute("INSERT INTO loan_payments (loan_id, payment_date, amount, principal_part, payment_type) VALUES (%s,'%s',%s,%s,'early_full')" % (lid, pd, amt, cb))
                audit_log(cur, staff, 'early_repayment', 'loan', lid, '', 'Полное досрочное погашение: %s' % amt, ip)
                conn.commit()
                return {'success': True, 'new_balance': 0, 'status': 'closed'}

            cur.execute("SELECT COUNT(*) FROM loan_schedule WHERE loan_id=%s AND status IN ('pending','partial','overdue')" % lid)
            remaining_periods = cur.fetchone()[0]
            cur.execute("DELETE FROM loan_schedule WHERE loan_id=%s AND status IN ('pending','partial','overdue')" % lid)
            cur.execute("SELECT COUNT(*) FROM loan_schedule WHERE loan_id=%s AND status='paid'" % lid)
            paid_count = cur.fetchone()[0]

            fn = calc_annuity_schedule if st == 'annuity' else calc_end_of_term_schedule

            if rt == 'reduce_payment':
                nt = max(remaining_periods, 1)
            else:
                if old_monthly > 0:
                    best_term = remaining_periods
                    for t in range(1, remaining_periods + 1):
                        _, m = fn(nb, r, t, date.fromisoformat(pd))
                        if m <= old_monthly * 1.1:
                            best_term = t
                            break
                    if best_term >= remaining_periods:
                        best_term = max(remaining_periods - 1, 1)
                    nt = max(best_term, 1)
                else:
                    nt = max(remaining_periods, 1)

            ns, nm = fn(nb, r, nt, date.fromisoformat(pd))
            for item in ns:
                cur.execute("INSERT INTO loan_schedule (loan_id,payment_no,payment_date,payment_amount,principal_amount,interest_amount,balance_after) VALUES (%s,%s,'%s',%s,%s,%s,%s)" % (lid, paid_count + item['payment_no'], item['payment_date'], item['payment_amount'], item['principal_amount'], item['interest_amount'], item['balance_after']))

            ne = date.fromisoformat(ns[-1]['payment_date'])
            total_term = paid_count + len(ns)
            cur.execute("UPDATE loans SET balance=%s, monthly_payment=%s, end_date='%s', term_months=%s, updated_at=NOW() WHERE id=%s" % (nb, nm, ne.isoformat(), total_term, lid))
            cur.execute("INSERT INTO loan_payments (loan_id,payment_date,amount,principal_part,payment_type) VALUES (%s,'%s',%s,%s,'early_partial')" % (lid, pd, amt, amt))
            refresh_loan_overdue_status(cur, lid)
            audit_log(cur, staff, 'early_repayment', 'loan', lid, '', 'Частичное: %s, тип: %s' % (amt, rt), ip)
            conn.commit()
            return {'success': True, 'new_balance': nb, 'new_schedule': ns, 'new_monthly': nm}

        elif action == 'update_payment':
            pid = int(body['payment_id'])
            cur.execute("SELECT loan_id, amount, principal_part, interest_part, penalty_part, payment_date FROM loan_payments WHERE id=%s" % pid)
            old = cur.fetchone()
            if not old:
                return {'error': 'Платёж не найден'}
            lid = old[0]
            old_principal = Decimal(str(old[2]))
            new_date = body.get('payment_date', str(old[5]))
            new_amount = Decimal(str(body.get('amount', float(old[1]))))
            new_pp = Decimal(str(body.get('principal_part', float(old[2]))))
            new_ip = Decimal(str(body.get('interest_part', float(old[3]))))
            new_pnp = Decimal(str(body.get('penalty_part', float(old[4]))))
            cur.execute("UPDATE loan_payments SET payment_date='%s', amount=%s, principal_part=%s, interest_part=%s, penalty_part=%s WHERE id=%s" % (
                new_date, float(new_amount), float(new_pp), float(new_ip), float(new_pnp), pid))
            delta_principal = new_pp - old_principal
            if delta_principal != 0:
                cur.execute("UPDATE loans SET balance=balance-%s, updated_at=NOW() WHERE id=%s" % (float(delta_principal), lid))
                cur.execute("SELECT balance FROM loans WHERE id=%s" % lid)
                nb = Decimal(str(cur.fetchone()[0]))
                if nb <= 0:
                    cur.execute("UPDATE loans SET balance=0, status='closed', updated_at=NOW() WHERE id=%s" % lid)
            recalc_loan_schedule_statuses(cur, lid)
            audit_log(cur, staff, 'update_payment', 'loan', lid, '', 'Платёж #%s: сумма %s, ОД %s, %%: %s' % (pid, float(new_amount), float(new_pp), float(new_ip)), ip)
            conn.commit()
            return {'success': True}

        elif action == 'delete_payment':
            pid = int(body['payment_id'])
            cur.execute("SELECT loan_id, principal_part FROM loan_payments WHERE id=%s" % pid)
            old = cur.fetchone()
            if not old:
                return {'error': 'Платёж не найден'}
            lid, old_pp = old[0], Decimal(str(old[1]))
            cur.execute("DELETE FROM loan_payments WHERE id=%s" % pid)
            if old_pp > 0:
                cur.execute("UPDATE loans SET balance=balance+%s, updated_at=NOW() WHERE id=%s" % (float(old_pp), lid))
            recalc_loan_schedule_statuses(cur, lid)
            audit_log(cur, staff, 'delete_payment', 'loan', lid, '', 'Удалён платёж #%s (ОД: %s)' % (pid, float(old_pp)), ip)
            conn.commit()
            return {'success': True}

        elif action == 'delete_contract':
            lid = int(body['loan_id'])
            cur.execute("SELECT contract_no FROM loans WHERE id=%s" % lid)
            lr = cur.fetchone()
            if not lr:
                return {'error': 'Договор не найден'}
            cur.execute("DELETE FROM loan_payments WHERE loan_id=%s" % lid)
            cur.execute("DELETE FROM loan_schedule WHERE loan_id=%s" % lid)
            cur.execute("DELETE FROM loans WHERE id=%s" % lid)
            audit_log(cur, staff, 'delete_contract', 'loan', lid, lr[0], '', ip)
            conn.commit()
            return {'success': True}

        elif action == 'delete_all_payments':
            lid = int(body['loan_id'])
            cur.execute("SELECT amount, rate, term_months, start_date, schedule_type FROM loans WHERE id=%s" % lid)
            lr = cur.fetchone()
            if not lr:
                return {'error': 'Договор не найден'}
            orig_amount = float(lr[0])
            r = float(lr[1])
            orig_term = int(lr[2])
            sd = date.fromisoformat(str(lr[3]))
            st = lr[4]
            cur.execute("DELETE FROM loan_payments WHERE loan_id=%s" % lid)
            cur.execute("DELETE FROM loan_schedule WHERE loan_id=%s" % lid)
            fn = calc_annuity_schedule if st == 'annuity' else calc_end_of_term_schedule
            ns, m = fn(orig_amount, r, orig_term, sd)
            for item in ns:
                cur.execute("INSERT INTO loan_schedule (loan_id,payment_no,payment_date,payment_amount,principal_amount,interest_amount,balance_after) VALUES (%s,%s,'%s',%s,%s,%s,%s)" % (lid, item['payment_no'], item['payment_date'], item['payment_amount'], item['principal_amount'], item['interest_amount'], item['balance_after']))
            ne = date.fromisoformat(ns[-1]['payment_date'])
            cur.execute("UPDATE loans SET balance=%s, monthly_payment=%s, term_months=%s, end_date='%s', status='active', updated_at=NOW() WHERE id=%s" % (orig_amount, m, orig_term, ne.isoformat(), lid))
            audit_log(cur, staff, 'delete_all_payments', 'loan', lid, '', '', ip)
            conn.commit()
            return {'success': True}

        elif action == 'fix_schedule':
            lid = int(body['loan_id'])
            cur.execute("SELECT amount, balance, rate, term_months, start_date, schedule_type FROM loans WHERE id=%s" % lid)
            lr = cur.fetchone()
            if not lr:
                return {'error': 'Договор не найден'}
            orig_amount, bal = float(lr[0]), Decimal(str(lr[1]))
            r, orig_term = float(lr[2]), int(lr[3])
            sd, st = date.fromisoformat(str(lr[4])), lr[5]

            cur.execute("""
                SELECT payment_no, COUNT(*) as cnt FROM loan_schedule 
                WHERE loan_id=%s GROUP BY payment_no HAVING COUNT(*) > 1
            """ % lid)
            dups = cur.fetchall()
            removed = 0
            for dup in dups:
                pno = dup[0]
                cur.execute("""
                    SELECT id FROM loan_schedule 
                    WHERE loan_id=%s AND payment_no=%s 
                    ORDER BY id
                """ % (lid, pno))
                ids = [row[0] for row in cur.fetchall()]
                cur.execute("""
                    SELECT id FROM loan_schedule 
                    WHERE loan_id=%s AND payment_no=%s AND status='paid'
                    ORDER BY id DESC LIMIT 1
                """ % (lid, pno))
                paid_row = cur.fetchone()
                keep_id = paid_row[0] if paid_row else ids[-1]
                for sid in ids:
                    if sid != keep_id:
                        cur.execute("DELETE FROM loan_schedule WHERE id=%s" % sid)
                        removed += 1

            recalc_loan_schedule_statuses(cur, lid)

            cur.execute("SELECT COALESCE(SUM(principal_part),0) FROM loan_payments WHERE loan_id=%s" % lid)
            total_paid_principal = Decimal(str(cur.fetchone()[0]))
            real_balance = Decimal(str(orig_amount)) - total_paid_principal
            if real_balance < 0:
                real_balance = Decimal('0')
            cur.execute("UPDATE loans SET balance=%s, updated_at=NOW() WHERE id=%s" % (float(real_balance), lid))

            if real_balance == 0:
                cur.execute("UPDATE loans SET status='closed', updated_at=NOW() WHERE id=%s" % lid)
            else:
                # Пересоздаём pending-строки графика от правильного баланса
                # Берём последний оплаченный период как точку отсчёта
                cur.execute("""
                    SELECT payment_date, payment_no FROM loan_schedule
                    WHERE loan_id=%s AND status='paid'
                    ORDER BY payment_no DESC LIMIT 1
                """ % lid)
                last_paid = cur.fetchone()

                cur.execute("SELECT COUNT(*) FROM loan_schedule WHERE loan_id=%s AND status IN ('pending','partial','overdue')" % lid)
                pending_count = cur.fetchone()[0]

                if last_paid and pending_count > 0 and real_balance > 0:
                    last_paid_date = last_paid[0] if isinstance(last_paid[0], date) else date.fromisoformat(str(last_paid[0]))
                    last_paid_no = last_paid[1]

                    cur.execute("UPDATE loan_schedule SET status='deleted_fix' WHERE loan_id=%s AND status IN ('pending','partial','overdue')" % lid)

                    fn = calc_annuity_schedule if st == 'annuity' else calc_end_of_term_schedule
                    new_sched, new_monthly = fn(float(real_balance), r, pending_count, last_paid_date)

                    for item in new_sched:
                        cur.execute("INSERT INTO loan_schedule (loan_id,payment_no,payment_date,payment_amount,principal_amount,interest_amount,balance_after) VALUES (%s,%s,'%s',%s,%s,%s,%s)" % (
                            lid, last_paid_no + item['payment_no'], item['payment_date'],
                            item['payment_amount'], item['principal_amount'],
                            item['interest_amount'], item['balance_after']))

                    ne = date.fromisoformat(new_sched[-1]['payment_date'])
                    new_term = last_paid_no + len(new_sched)
                    cur.execute("UPDATE loans SET monthly_payment=%s, end_date='%s', term_months=%s, updated_at=NOW() WHERE id=%s" % (
                        new_monthly, ne.isoformat(), new_term, lid))
                    removed += pending_count

                refresh_loan_overdue_status(cur, lid)

            audit_log(cur, staff, 'fix_schedule', 'loan', lid, '', 'Удалено дублей: %s, пересчитан баланс: %s' % (removed, float(real_balance)), ip)
            conn.commit()
            return {'success': True, 'removed_duplicates': removed, 'new_balance': float(real_balance)}

        elif action == 'recalc_statuses':
            lid = int(body['loan_id'])
            recalc_loan_schedule_statuses(cur, lid)
            audit_log(cur, staff, 'recalc_statuses', 'loan', lid, '', 'Пересчёт статусов платежей', ip)
            conn.commit()
            return {'success': True}

        elif action == 'rebuild_schedule':
            """Пересоздаёт график с оригинальной даты начала, сохраняя платежи"""
            lid = int(body['loan_id'])
            cur.execute("SELECT amount, rate, term_months, start_date, schedule_type FROM loans WHERE id=%s" % lid)
            lr = cur.fetchone()
            if not lr:
                return {'error': 'Договор не найден'}
            orig_amount = float(lr[0])
            r = safe_float(body['rate'], 'ставка') if body.get('rate') else float(lr[1])
            orig_term = safe_int(body['term_months'], 'срок') if body.get('term_months') else int(lr[2])
            sd, st = date.fromisoformat(str(lr[3])), lr[4]
            cur.execute("DELETE FROM loan_schedule WHERE loan_id=%s" % lid)
            fn = calc_annuity_schedule if st == 'annuity' else calc_end_of_term_schedule
            ns, m = fn(orig_amount, r, orig_term, sd)
            for item in ns:
                cur.execute("INSERT INTO loan_schedule (loan_id,payment_no,payment_date,payment_amount,principal_amount,interest_amount,balance_after) VALUES (%s,%s,'%s',%s,%s,%s,%s)" % (lid, item['payment_no'], item['payment_date'], item['payment_amount'], item['principal_amount'], item['interest_amount'], item['balance_after']))
            ne = date.fromisoformat(ns[-1]['payment_date'])
            cur.execute("UPDATE loans SET rate=%s, term_months=%s, monthly_payment=%s, end_date='%s', updated_at=NOW() WHERE id=%s" % (r, orig_term, m, ne.isoformat(), lid))
            recalc_loan_schedule_statuses(cur, lid)
            audit_log(cur, staff, 'rebuild_schedule', 'loan', lid, '', 'Пересоздан график с %s, %s периодов' % (sd.isoformat(), orig_term), ip)
            conn.commit()
            return {'success': True, 'periods': len(ns), 'monthly_payment': m, 'end_date': ne.isoformat()}

        elif action == 'modify':
            lid = int(body['loan_id'])
            cur.execute("SELECT balance, rate, term_months, start_date, schedule_type FROM loans WHERE id=%s" % lid)
            lr = cur.fetchone()
            bal = float(lr[0])
            r = safe_float(body['new_rate'], 'ставка') if body.get('new_rate') else float(lr[1])
            t = safe_int(body['new_term'], 'срок') if body.get('new_term') else int(lr[2])
            st = lr[4]

            cur.execute("DELETE FROM loan_schedule WHERE loan_id=%s AND status IN ('pending','partial','overdue')" % lid)
            cur.execute("SELECT MAX(payment_no) FROM loan_schedule WHERE loan_id=%s AND status='paid'" % lid)
            max_paid_no_row = cur.fetchone()
            max_paid_no = max_paid_no_row[0] if max_paid_no_row and max_paid_no_row[0] else 0
            fn = calc_annuity_schedule if st == 'annuity' else calc_end_of_term_schedule
            ns, m = fn(bal, r, t, date.today())
            for item in ns:
                cur.execute("INSERT INTO loan_schedule (loan_id,payment_no,payment_date,payment_amount,principal_amount,interest_amount,balance_after) VALUES (%s,%s,'%s',%s,%s,%s,%s)" % (lid, max_paid_no + item['payment_no'], item['payment_date'], item['payment_amount'], item['principal_amount'], item['interest_amount'], item['balance_after']))
            ne = date.fromisoformat(ns[-1]['payment_date'])
            total_term = max_paid_no + len(ns)
            cur.execute("UPDATE loans SET rate=%s, term_months=%s, monthly_payment=%s, end_date='%s', updated_at=NOW() WHERE id=%s" % (r, total_term, m, ne.isoformat(), lid))
            audit_log(cur, staff, 'modify', 'loan', lid, '', 'Ставка: %s%%, срок: %s мес.' % (r, t), ip)
            conn.commit()
            return {'success': True, 'new_schedule': ns, 'monthly_payment': m}

def calc_savings_schedule_with_transactions(initial_amount, rate, term, start_date, payout_type, transactions, rate_changes=None):
    base_rate = Decimal(str(rate))
    rc_list = []
    if rate_changes:
        for rc in rate_changes:
            rc_date = date.fromisoformat(str(rc[0])) if not isinstance(rc[0], date) else rc[0]
            rc_list.append((rc_date, Decimal(str(rc[1]))))
        rc_list.sort(key=lambda x: x[0])
    schedule = []
    cumulative = Decimal('0')
    bal_changes = []
    bal_changes.append((start_date, Decimal(str(initial_amount))))
    for tx in transactions:
        tx_date = date.fromisoformat(str(tx[0])) if not isinstance(tx[0], date) else tx[0]
        tx_amt = Decimal(str(tx[1]))
        tx_type = tx[2]
        if tx_type == 'deposit':
            bal_changes.append((tx_date, tx_amt))
        elif tx_type in ('withdrawal', 'partial_withdrawal'):
            bal_changes.append((tx_date, -tx_amt))
    bal_changes.sort(key=lambda x: x[0])

    def get_rate_on_date(d):
        r = base_rate
        for rc_d, rc_r in rc_list:
            if rc_d <= d:
                r = rc_r
        return r

    for i in range(1, term + 1):
        period_start = last_day_of_month(add_months(start_date, i - 2)) if i > 1 else start_date
        period_end = last_day_of_month(add_months(start_date, i - 1))
        running_bal = Decimal('0')
        for bd, ba in bal_changes:
            if bd <= period_start:
                running_bal += ba
        interest = Decimal('0')
        split_dates = set()
        split_dates.add(period_start)
        for bd, ba in bal_changes:
            if period_start < bd <= period_end:
                split_dates.add(bd)
        for rc_d, _ in rc_list:
            if period_start < rc_d <= period_end:
                split_dates.add(rc_d)
        split_dates = sorted(split_dates)
        current_bal = running_bal
        for j, ss in enumerate(split_dates):
            se = split_dates[j + 1] if j + 1 < len(split_dates) else period_end
            days = (se - ss).days
            if j == len(split_dates) - 1:
                days = (period_end - ss).days
            r = get_rate_on_date(ss)
            if days > 0 and current_bal > 0:
                day_interest = (current_bal * r / Decimal('100') * Decimal(str(days)) / Decimal('365')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                interest += day_interest
            if j + 1 < len(split_dates):
                for bd, ba in bal_changes:
                    if bd == split_dates[j + 1]:
                        current_bal += ba
        cumulative += interest
        final_bal = Decimal('0')
        for bd, ba in bal_changes:
            if bd <= period_end:
                final_bal += ba
        current_rate = float(get_rate_on_date(period_end))
        balance_after = float(final_bal + cumulative) if payout_type == 'end_of_term' else float(final_bal)
        schedule.append({
            'period_no': i, 'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat(), 'interest_amount': float(interest),
            'cumulative_interest': float(cumulative), 'balance_after': balance_after,
            'rate': current_rate,
        })
    return schedule

def recalc_savings_schedule(cur, sid, amount, rate, term, start_date, payout_type):
    cur.execute("DELETE FROM savings_schedule WHERE saving_id=%s AND status IN ('pending','accrued')" % sid)
    cur.execute("SELECT transaction_date, amount, transaction_type FROM savings_transactions WHERE saving_id=%s AND transaction_type IN ('deposit','withdrawal','partial_withdrawal') ORDER BY transaction_date, id" % sid)
    transactions = cur.fetchall()
    cur.execute("SELECT amount FROM savings WHERE id=%s" % sid)
    sv_row = cur.fetchone()
    initial_amount = float(sv_row[0])
    for tx in transactions:
        tx_type = tx[2]
        tx_amt = float(tx[1])
        if tx_type == 'deposit':
            initial_amount -= tx_amt
        elif tx_type in ('withdrawal', 'partial_withdrawal'):
            initial_amount += tx_amt
    if initial_amount < 0:
        initial_amount = 0
    cur.execute("SELECT effective_date, new_rate FROM savings_rate_changes WHERE saving_id=%s ORDER BY effective_date" % sid)
    rate_changes = cur.fetchall()
    schedule = calc_savings_schedule_with_transactions(initial_amount, rate, term, start_date, payout_type, transactions, rate_changes)
    today = date.today().isoformat()
    for item in schedule:
        cur.execute("SELECT id FROM savings_schedule WHERE saving_id=%s AND period_no=%s AND status='paid'" % (sid, item['period_no']))
        paid_row = cur.fetchone()
        if paid_row:
            cur.execute("UPDATE savings_schedule SET period_start='%s', period_end='%s', interest_amount=%s, cumulative_interest=%s, balance_after=%s WHERE id=%s" % (
                item['period_start'], item['period_end'], item['interest_amount'], item['cumulative_interest'], item['balance_after'], paid_row[0]))
            continue
        new_status = 'accrued' if item['period_end'] <= today else 'pending'
        cur.execute("INSERT INTO savings_schedule (saving_id,period_no,period_start,period_end,interest_amount,cumulative_interest,balance_after,status) VALUES (%s,%s,'%s','%s',%s,%s,%s,'%s')" % (
            sid, item['period_no'], item['period_start'], item['period_end'], item['interest_amount'], item['cumulative_interest'], item['balance_after'], new_status))
    return schedule

def get_accrued_interest_end_of_prev_month(cur, sid):
    today = date.today()
    first_of_month = today.replace(day=1)
    last_of_prev = first_of_month - timedelta(days=1)
    cur.execute("SELECT COALESCE(SUM(daily_amount), 0) FROM savings_daily_accruals WHERE saving_id=%s AND accrual_date <= '%s'" % (sid, last_of_prev.isoformat()))
    total_accrued = Decimal(str(cur.fetchone()[0]))
    cur.execute("SELECT paid_interest FROM savings WHERE id=%s" % sid)
    paid = Decimal(str(cur.fetchone()[0]))
    available = total_accrued - paid
    return max(available, Decimal('0'))

def handle_savings(method, params, body, cur, conn, staff=None, ip=''):
    if method == 'GET':
        action = params.get('action', 'list')
        if action == 'detail':
            s = query_one(cur, "SELECT * FROM savings WHERE id=%s" % params['id'])
            if not s: return None
            cur.execute("SELECT CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name) ELSE m.company_name END FROM members m WHERE m.id=%s" % s['member_id'])
            nr = cur.fetchone()
            s['member_name'] = nr[0] if nr else ''
            if s.get('org_id'):
                org_row = query_one(cur, "SELECT name, short_name FROM organizations WHERE id=%s" % s['org_id'])
                s['org_name'] = org_row['name'] if org_row else ''
                s['org_short_name'] = org_row['short_name'] if org_row else ''
            today = date.today().isoformat()
            cur.execute("UPDATE savings_schedule SET status='accrued' WHERE saving_id=%s AND status='pending' AND period_end <= '%s'" % (params['id'], today))
            if cur.rowcount > 0:
                conn.commit()
            s['schedule'] = query_rows(cur, "SELECT * FROM savings_schedule WHERE saving_id=%s ORDER BY period_no" % params['id'])
            s['transactions'] = query_rows(cur, "SELECT * FROM savings_transactions WHERE saving_id=%s ORDER BY transaction_date" % params['id'])
            s['daily_accruals'] = query_rows(cur, "SELECT id, accrual_date, balance, rate, daily_amount, created_at FROM savings_daily_accruals WHERE saving_id=%s ORDER BY accrual_date" % params['id'])
            s['rate_changes'] = query_rows(cur, "SELECT id, effective_date, old_rate, new_rate, reason, created_at FROM savings_rate_changes WHERE saving_id=%s ORDER BY effective_date" % params['id'])
            cur.execute("SELECT COALESCE(SUM(daily_amount), 0) FROM savings_daily_accruals WHERE saving_id=%s" % params['id'])
            s['total_daily_accrued'] = float(cur.fetchone()[0])
            s['max_payout'] = float(get_accrued_interest_end_of_prev_month(cur, int(params['id'])))
            cur.execute("SELECT MIN(accrual_date), MAX(accrual_date), COUNT(*) FROM savings_daily_accruals WHERE saving_id=%s" % params['id'])
            accrual_info = cur.fetchone()
            s['accrual_first_date'] = str(accrual_info[0]) if accrual_info[0] else None
            s['accrual_last_date'] = str(accrual_info[1]) if accrual_info[1] else None
            s['accrual_days_count'] = accrual_info[2] or 0
            return s
        elif action == 'schedule':
            a, r, t = safe_float(params['amount'], 'сумма'), safe_float(params['rate'], 'ставка'), safe_int(params['term'], 'срок')
            pt = params.get('payout_type', 'monthly')
            sd = date.fromisoformat(params.get('start_date', date.today().isoformat()))
            return {'schedule': calc_savings_schedule(a, r, t, sd, pt)}
        else:
            return query_rows(cur, """
                SELECT s.id, s.contract_no, s.amount, s.rate, s.term_months, s.payout_type,
                       s.start_date, s.end_date, s.accrued_interest, s.paid_interest, s.current_balance, s.status,
                       s.min_balance_pct, s.org_id,
                       CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name)
                            ELSE m.company_name END as member_name, m.id as member_id,
                       o.name as org_name, o.short_name as org_short_name
                FROM savings s JOIN members m ON m.id=s.member_id
                LEFT JOIN organizations o ON o.id=s.org_id
                ORDER BY s.created_at DESC
            """)

    elif method == 'POST':
        action = body.get('action', 'create')
        if action == 'create':
            cn = body.get('contract_no', '').strip()
            mid = int(body['member_id'])
            a, r, t = safe_float(body['amount'], 'сумма'), safe_float(body['rate'], 'ставка'), safe_int(body['term_months'], 'срок')
            pt = body.get('payout_type', 'monthly')
            sd = date.fromisoformat(body.get('start_date', date.today().isoformat()))
            mbp = safe_float(body.get('min_balance_pct', 0), 'мин. остаток %')
            org_id = body.get('org_id')
            if not cn:
                cur.execute("SELECT MAX(CAST(SUBSTRING(contract_no FROM '^[0-9]+') AS INTEGER)) FROM savings WHERE contract_no ~ '^[0-9]+'")
                max_num = cur.fetchone()[0] or 0
                cn = '%s-%s' % (max_num + 1, sd.strftime('%d%m%Y'))
            else:
                cur.execute("SELECT id FROM savings WHERE contract_no='%s'" % esc(cn))
                if cur.fetchone():
                    return {'statusCode': 400, 'headers': cors, 'body': json.dumps({'error': 'Договор с номером %s уже существует' % cn})}
            schedule = calc_savings_schedule(a, r, t, sd, pt)
            ed = add_months(sd, t)
            cur.execute("INSERT INTO savings (contract_no,member_id,amount,rate,term_months,payout_type,start_date,end_date,current_balance,status,min_balance_pct,org_id) VALUES ('%s',%s,%s,%s,%s,'%s','%s','%s',%s,'active',%s,%s) RETURNING id" % (esc(cn), mid, a, r, t, pt, sd.isoformat(), ed.isoformat(), a, mbp, org_id if org_id else 'NULL'))
            sid = cur.fetchone()[0]
            for item in schedule:
                cur.execute("INSERT INTO savings_schedule (saving_id,period_no,period_start,period_end,interest_amount,cumulative_interest,balance_after) VALUES (%s,%s,'%s','%s',%s,%s,%s)" % (sid, item['period_no'], item['period_start'], item['period_end'], item['interest_amount'], item['cumulative_interest'], item['balance_after']))
            cur.execute("INSERT INTO savings_transactions (saving_id,transaction_date,amount,transaction_type,description) VALUES (%s,'%s',%s,'opening','Открытие договора. Сумма: %s руб., ставка: %s%%, срок: %s мес.')" % (sid, sd.isoformat(), a, fmt_money(a), r, t))
            audit_log(cur, staff, 'create', 'saving', sid, cn, 'Сумма: %s, ставка: %s%%, срок: %s мес., несниж.остаток: %s%%' % (a, r, t, mbp), ip)
            conn.commit()
            return {'id': sid, 'contract_no': cn, 'schedule': schedule}

        elif action == 'transaction':
            sid = int(body['saving_id'])
            a = safe_float(body['amount'], 'сумма')
            tt = body['transaction_type']
            td = body.get('transaction_date', date.today().isoformat())
            ic = body.get('is_cash', False)
            d = body.get('description', '')
            cur.execute("INSERT INTO savings_transactions (saving_id,transaction_date,amount,transaction_type,is_cash,description) VALUES (%s,'%s',%s,'%s',%s,'%s')" % (sid, td, a, tt, ic, esc(d)))
            if tt == 'deposit':
                cur.execute("UPDATE savings SET current_balance=current_balance+%s, amount=amount+%s, updated_at=NOW() WHERE id=%s" % (a, a, sid))
                cur.execute("SELECT amount, rate, term_months, start_date, payout_type FROM savings WHERE id=%s" % sid)
                sv = cur.fetchone()
                recalc_savings_schedule(cur, sid, float(sv[0]), float(sv[1]), int(sv[2]), date.fromisoformat(str(sv[3])), sv[4])
            elif tt == 'withdrawal':
                cur.execute("UPDATE savings SET current_balance=current_balance-%s, updated_at=NOW() WHERE id=%s" % (a, sid))
            elif tt == 'interest_payout':
                cur.execute("UPDATE savings SET paid_interest=paid_interest+%s, updated_at=NOW() WHERE id=%s" % (a, sid))
            tt_labels = {'deposit': 'Пополнение', 'withdrawal': 'Снятие', 'interest_payout': 'Выплата %'}
            audit_log(cur, staff, 'transaction', 'saving', sid, '', '%s: %s' % (tt_labels.get(tt, tt), a), ip)
            conn.commit()
            return {'success': True}

        elif action == 'interest_payout':
            sid = int(body['saving_id'])
            cur.execute("SELECT payout_type, accrued_interest, paid_interest, current_balance, rate, amount FROM savings WHERE id=%s" % sid)
            sv = cur.fetchone()
            if not sv:
                return {'error': 'Вклад не найден'}
            td = body.get('transaction_date', date.today().isoformat())

            max_payout = get_accrued_interest_end_of_prev_month(cur, sid)
            if max_payout <= 0:
                return {'error': 'Нет начисленных процентов к выплате (по данным на конец предыдущего месяца)'}

            interest = Decimal(str(safe_float(body['amount'], 'сумма'))) if body.get('amount') else max_payout
            if interest > max_payout:
                interest = max_payout

            cur.execute("INSERT INTO savings_transactions (saving_id,transaction_date,amount,transaction_type,description) VALUES (%s,'%s',%s,'interest_payout','Выплата процентов')" % (sid, td, float(interest)))
            cur.execute("UPDATE savings SET paid_interest=paid_interest+%s, updated_at=NOW() WHERE id=%s" % (float(interest), sid))
            audit_log(cur, staff, 'interest_payout', 'saving', sid, '', 'Выплата %%: %s (макс: %s)' % (float(interest), float(max_payout)), ip)
            conn.commit()
            return {'success': True, 'amount': float(interest), 'max_payout': float(max_payout)}

        elif action == 'daily_accrue':
            today = date.today()
            accrual_date = body.get('date', today.isoformat())
            cur.execute("SELECT id, current_balance, rate, start_date FROM savings WHERE status='active'")
            savings_rows = cur.fetchall()
            count = 0
            total = Decimal('0')
            for row in savings_rows:
                s_id, s_bal, s_rate, s_start = row[0], Decimal(str(row[1])), Decimal(str(row[2])), str(row[3])
                if s_bal <= 0:
                    continue
                if accrual_date <= s_start:
                    continue
                daily_amount = (s_bal * s_rate / Decimal('100') / Decimal('365')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                if daily_amount <= 0:
                    continue
                cur.execute("SELECT id FROM savings_daily_accruals WHERE saving_id=%s AND accrual_date='%s'" % (s_id, accrual_date))
                if cur.fetchone():
                    continue
                cur.execute("INSERT INTO savings_daily_accruals (saving_id, accrual_date, balance, rate, daily_amount) VALUES (%s, '%s', %s, %s, %s)" % (s_id, accrual_date, float(s_bal), float(s_rate), float(daily_amount)))
                cur.execute("UPDATE savings SET accrued_interest=accrued_interest+%s, updated_at=NOW() WHERE id=%s" % (float(daily_amount), s_id))
                count += 1
                total += daily_amount
            if count > 0:
                conn.commit()
            return {'success': True, 'accrued_count': count, 'total_amount': float(total), 'date': accrual_date}

        elif action == 'backfill_accrue':
            sid = int(body['saving_id'])
            date_from = body.get('date_from')
            date_to = body.get('date_to', date.today().isoformat())
            if not date_from:
                cur.execute("SELECT start_date FROM savings WHERE id=%s" % sid)
                sr = cur.fetchone()
                if not sr:
                    return {'error': 'Вклад не найден'}
                date_from = str(sr[0])

            cur.execute("SELECT id, current_balance, rate, start_date FROM savings WHERE id=%s AND status='active'" % sid)
            sv = cur.fetchone()
            if not sv:
                return {'error': 'Вклад не найден или не активен'}

            s_bal_current = Decimal(str(sv[1]))
            s_rate = Decimal(str(sv[2]))
            s_start = str(sv[3])
            cur.execute("SELECT effective_date, new_rate FROM savings_rate_changes WHERE saving_id=%s ORDER BY effective_date" % sid)
            rate_changes_rows = cur.fetchall()
            rate_changes_list = [(date.fromisoformat(str(r[0])), Decimal(str(r[1]))) for r in rate_changes_rows]

            s_start_date = date.fromisoformat(s_start)
            d = date.fromisoformat(date_from)
            d_to = date.fromisoformat(date_to)
            if d <= s_start_date:
                d = s_start_date + timedelta(days=1)

            cur.execute("""
                SELECT accrual_date, balance FROM savings_daily_accruals 
                WHERE saving_id=%s ORDER BY accrual_date
            """ % sid)
            existing = {str(r[0]): Decimal(str(r[1])) for r in cur.fetchall()}

            cur.execute("""
                SELECT transaction_date, amount, transaction_type FROM savings_transactions 
                WHERE saving_id=%s ORDER BY transaction_date, id
            """ % sid)
            tx_rows = cur.fetchall()
            tx_by_date = {}
            for tr in tx_rows:
                td_str = str(tr[0])
                if td_str not in tx_by_date:
                    tx_by_date[td_str] = []
                tx_by_date[td_str].append((Decimal(str(tr[1])), tr[2]))

            cur.execute("SELECT amount FROM savings WHERE id=%s" % sid)
            initial_amount = Decimal(str(cur.fetchone()[0]))

            running_bal = initial_amount
            for td_str in sorted(tx_by_date.keys()):
                if td_str < d.isoformat():
                    for amt, tt in tx_by_date[td_str]:
                        if tt == 'deposit':
                            running_bal += amt
                        elif tt == 'withdrawal':
                            running_bal -= amt

            count = 0
            total = Decimal('0')
            while d <= d_to:
                ds = d.isoformat()
                if running_bal > 0 and ds not in existing:
                    effective_rate = s_rate
                    for rc_d, rc_r in rate_changes_list:
                        if rc_d <= d:
                            effective_rate = rc_r
                    daily_amount = (running_bal * effective_rate / Decimal('100') / Decimal('365')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    if daily_amount > 0:
                        cur.execute("INSERT INTO savings_daily_accruals (saving_id, accrual_date, balance, rate, daily_amount) VALUES (%s, '%s', %s, %s, %s)" % (sid, ds, float(running_bal), float(effective_rate), float(daily_amount)))
                        cur.execute("UPDATE savings SET accrued_interest=accrued_interest+%s, updated_at=NOW() WHERE id=%s" % (float(daily_amount), sid))
                        count += 1
                        total += daily_amount

                if ds in tx_by_date:
                    for amt, tt in tx_by_date[ds]:
                        if tt == 'deposit':
                            running_bal += amt
                        elif tt == 'withdrawal':
                            running_bal -= amt

                d += timedelta(days=1)

            if count > 0:
                cur.execute("INSERT INTO savings_transactions (saving_id,transaction_date,amount,transaction_type,description) VALUES (%s,'%s',%s,'interest_accrual','Начисление процентов за %s — %s (%s дн.)')" % (sid, date_to, float(total), fmt_date(date_from), fmt_date(date_to), count))
            audit_log(cur, staff, 'backfill_accrue', 'saving', sid, '', 'Период: %s — %s, дней: %s, сумма: %s' % (date_from, date_to, count, float(total)), ip)
            if count > 0:
                conn.commit()
            return {'success': True, 'days_accrued': count, 'total_amount': float(total), 'date_from': date_from, 'date_to': date_to}

        elif action == 'partial_withdrawal':
            sid = int(body['saving_id'])
            a = Decimal(str(safe_float(body['amount'], 'сумма')))
            td = body.get('transaction_date', date.today().isoformat())
            cur.execute("SELECT current_balance, amount, min_balance_pct FROM savings WHERE id=%s AND status='active'" % sid)
            sv = cur.fetchone()
            if not sv:
                return {'error': 'Вклад не найден или не активен'}
            bal = Decimal(str(sv[0]))
            orig_amount = Decimal(str(sv[1]))
            mbp = Decimal(str(sv[2]))
            if mbp <= 0:
                return {'error': 'По данному договору частичное изъятие не предусмотрено'}
            min_bal = (orig_amount * mbp / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            new_bal = bal - a
            if new_bal < min_bal:
                return {'error': 'Нельзя снять больше. Неснижаемый остаток: %s руб. (текущий баланс: %s руб.)' % (float(min_bal), float(bal))}
            cur.execute("INSERT INTO savings_transactions (saving_id,transaction_date,amount,transaction_type,description) VALUES (%s,'%s',%s,'partial_withdrawal','Частичное изъятие')" % (sid, td, float(a)))
            cur.execute("UPDATE savings SET current_balance=current_balance-%s, amount=amount-%s, updated_at=NOW() WHERE id=%s" % (float(a), float(a), sid))
            cur.execute("SELECT amount, rate, term_months, start_date, payout_type FROM savings WHERE id=%s" % sid)
            sv2 = cur.fetchone()
            recalc_savings_schedule(cur, sid, float(sv2[0]), float(sv2[1]), int(sv2[2]), date.fromisoformat(str(sv2[3])), sv2[4])
            audit_log(cur, staff, 'partial_withdrawal', 'saving', sid, '', 'Сумма: %s, остаток: %s' % (float(a), float(new_bal)), ip)
            conn.commit()
            return {'success': True, 'new_balance': float(new_bal), 'min_balance': float(min_bal)}

        elif action == 'modify_term':
            sid = int(body['saving_id'])
            new_term = safe_int(body.get('new_term'), 'срок')
            effective_date = body.get('effective_date', date.today().isoformat())
            cur.execute("SELECT amount, rate, start_date, payout_type FROM savings WHERE id=%s AND status='active'" % sid)
            sv = cur.fetchone()
            if not sv:
                return {'error': 'Вклад не найден или не активен'}
            s_amount, s_rate, s_start, s_pt = float(sv[0]), float(sv[1]), date.fromisoformat(str(sv[2])), sv[3]
            schedule = recalc_savings_schedule(cur, sid, s_amount, s_rate, new_term, s_start, s_pt)
            new_end = add_months(s_start, new_term)
            cur.execute("UPDATE savings SET term_months=%s, end_date='%s', updated_at=NOW() WHERE id=%s" % (new_term, new_end.isoformat(), sid))
            cur.execute("INSERT INTO savings_transactions (saving_id,transaction_date,amount,transaction_type,description) VALUES (%s,'%s',0,'term_change','Изменение срока: %s мес. с %s, новая дата окончания: %s')" % (sid, effective_date, new_term, fmt_date(effective_date), fmt_date(new_end.isoformat())))
            audit_log(cur, staff, 'modify_term', 'saving', sid, '', 'Новый срок: %s мес. с %s' % (new_term, effective_date), ip)
            conn.commit()
            return {'success': True, 'new_term': new_term, 'new_end_date': new_end.isoformat(), 'schedule': schedule}

        elif action == 'recalc_schedule':
            sid = int(body['saving_id'])
            cur.execute("SELECT amount, rate, term_months, start_date, payout_type FROM savings WHERE id=%s AND status='active'" % sid)
            sv = cur.fetchone()
            if not sv:
                return {'error': 'Вклад не найден или не активен'}
            s_amount, s_rate, s_term = float(sv[0]), float(sv[1]), int(sv[2])
            s_start, s_pt = date.fromisoformat(str(sv[3])), sv[4]
            schedule = recalc_savings_schedule(cur, sid, s_amount, s_rate, s_term, s_start, s_pt)
            new_end = add_months(s_start, s_term)
            cur.execute("UPDATE savings SET end_date='%s', updated_at=NOW() WHERE id=%s" % (new_end.isoformat(), sid))
            audit_log(cur, staff, 'recalc_schedule', 'saving', sid, '', 'Пересчёт графика', ip)
            conn.commit()
            return {'success': True, 'schedule': schedule, 'new_end_date': new_end.isoformat()}
        
        elif action == 'recalc_all_active':
            cur.execute("SELECT id, amount, rate, term_months, start_date, payout_type, contract_no FROM savings WHERE status='active'")
            active_savings = cur.fetchall()
            recalculated = 0
            errors = []
            for sv in active_savings:
                sid, s_amount, s_rate, s_term, s_start_str, s_pt, cn = sv[0], float(sv[1]), float(sv[2]), int(sv[3]), str(sv[4]), sv[5], sv[6]
                s_start = date.fromisoformat(s_start_str)
                try:
                    recalc_savings_schedule(cur, sid, s_amount, s_rate, s_term, s_start, s_pt)
                    new_end = add_months(s_start, s_term)
                    cur.execute("UPDATE savings SET end_date='%s', updated_at=NOW() WHERE id=%s" % (new_end.isoformat(), sid))
                    audit_log(cur, staff, 'recalc_schedule', 'saving', sid, cn, 'Массовый пересчёт графика', ip)
                    recalculated += 1
                except Exception as e:
                    errors.append({'contract_no': cn, 'error': str(e)})
            conn.commit()
            return {'success': True, 'recalculated': recalculated, 'total': len(active_savings), 'errors': errors}

        elif action == 'change_rate':
            sid = int(body['saving_id'])
            new_rate = safe_float(body.get('new_rate'), 'ставка')
            effective_date = body.get('effective_date', date.today().isoformat())
            reason = body.get('reason', '')
            cur.execute("SELECT rate, amount, term_months, start_date, payout_type FROM savings WHERE id=%s AND status='active'" % sid)
            sv = cur.fetchone()
            if not sv:
                return {'error': 'Вклад не найден или не активен'}
            old_rate = float(sv[0])
            if new_rate == old_rate:
                return {'error': 'Новая ставка совпадает с текущей'}
            cur.execute("INSERT INTO savings_rate_changes (saving_id, effective_date, old_rate, new_rate, reason, created_by) VALUES (%s, '%s', %s, %s, '%s', %s)" % (sid, effective_date, old_rate, new_rate, esc(reason), staff))
            cur.execute("UPDATE savings SET rate=%s, updated_at=NOW() WHERE id=%s" % (new_rate, sid))
            s_amount, s_term = float(sv[1]), int(sv[2])
            s_start, s_pt = date.fromisoformat(str(sv[3])), sv[4]
            schedule = recalc_savings_schedule(cur, sid, s_amount, new_rate, s_term, s_start, s_pt)
            cur.execute("INSERT INTO savings_transactions (saving_id,transaction_date,amount,transaction_type,description) VALUES (%s,'%s',0,'rate_change','Изменение ставки: %s%% → %s%% с %s. %s')" % (sid, effective_date, old_rate, new_rate, fmt_date(effective_date), esc(reason)))
            audit_log(cur, staff, 'change_rate', 'saving', sid, '', 'Ставка: %s%% → %s%% с %s' % (old_rate, new_rate, effective_date), ip)
            conn.commit()
            return {'success': True, 'old_rate': old_rate, 'new_rate': new_rate, 'schedule': schedule}

        elif action == 'auto_accrue':
            sid = int(body['saving_id'])
            today = date.today()
            cur.execute("SELECT id, period_no, period_end FROM savings_schedule WHERE saving_id=%s AND status='pending' AND period_end <= '%s' ORDER BY period_no" % (sid, today.isoformat()))
            rows = cur.fetchall()
            count = 0
            for r in rows:
                cur.execute("UPDATE savings_schedule SET status='accrued' WHERE id=%s" % r[0])
                count += 1
            if count > 0:
                conn.commit()
            return {'success': True, 'accrued_count': count}

        elif action == 'update_transaction':
            tid = int(body['transaction_id'])
            cur.execute("SELECT saving_id, amount, transaction_type FROM savings_transactions WHERE id=%s" % tid)
            old = cur.fetchone()
            if not old:
                return {'error': 'Операция не найдена'}
            sid, old_amount, old_tt = old[0], Decimal(str(old[1])), old[2]

            new_date = body.get('transaction_date')
            new_amount = Decimal(str(body.get('amount', float(old_amount))))
            new_desc = body.get('description', '')
            delta = new_amount - old_amount

            upd = "amount=%s" % float(new_amount)
            if new_date:
                upd += ", transaction_date='%s'" % new_date
            if new_desc is not None:
                upd += ", description='%s'" % esc(new_desc)
            cur.execute("UPDATE savings_transactions SET %s WHERE id=%s" % (upd, tid))

            if delta != 0:
                if old_tt == 'deposit':
                    cur.execute("UPDATE savings SET current_balance=current_balance+%s, amount=amount+%s, updated_at=NOW() WHERE id=%s" % (float(delta), float(delta), sid))
                elif old_tt == 'withdrawal':
                    cur.execute("UPDATE savings SET current_balance=current_balance-%s, updated_at=NOW() WHERE id=%s" % (float(delta), sid))
                elif old_tt == 'interest_payout':
                    cur.execute("UPDATE savings SET paid_interest=paid_interest+%s, accrued_interest=accrued_interest+%s, updated_at=NOW() WHERE id=%s" % (float(delta), float(delta), sid))
            audit_log(cur, staff, 'update_transaction', 'saving', sid, '', 'Транзакция #%s: %s -> %s' % (tid, float(old_amount), float(new_amount)), ip)
            conn.commit()
            return {'success': True}

        elif action == 'delete_transaction':
            tid = int(body['transaction_id'])
            cur.execute("SELECT saving_id, amount, transaction_type FROM savings_transactions WHERE id=%s" % tid)
            old = cur.fetchone()
            if not old:
                return {'error': 'Операция не найдена'}
            sid, old_amount, old_tt = old[0], Decimal(str(old[1])), old[2]
            cur.execute("DELETE FROM savings_transactions WHERE id=%s" % tid)
            if old_tt == 'deposit':
                cur.execute("UPDATE savings SET current_balance=current_balance-%s, amount=amount-%s, updated_at=NOW() WHERE id=%s" % (float(old_amount), float(old_amount), sid))
            elif old_tt == 'withdrawal':
                cur.execute("UPDATE savings SET current_balance=current_balance+%s, updated_at=NOW() WHERE id=%s" % (float(old_amount), sid))
            elif old_tt == 'interest_payout':
                cur.execute("UPDATE savings SET paid_interest=paid_interest-%s, accrued_interest=accrued_interest-%s, updated_at=NOW() WHERE id=%s" % (float(old_amount), float(old_amount), sid))
            audit_log(cur, staff, 'delete_transaction', 'saving', sid, '', 'Удалена: %s %s' % (old_tt, float(old_amount)), ip)
            conn.commit()
            return {'success': True}

        elif action == 'delete_contract':
            sid = int(body['saving_id'])
            cur.execute("SELECT contract_no FROM savings WHERE id=%s" % sid)
            sr = cur.fetchone()
            if not sr:
                return {'error': 'Договор не найден'}
            cur.execute("DELETE FROM savings_transactions WHERE saving_id=%s" % sid)
            cur.execute("DELETE FROM savings_schedule WHERE saving_id=%s" % sid)
            cur.execute("DELETE FROM savings_rate_changes WHERE saving_id=%s" % sid)
            cur.execute("DELETE FROM savings WHERE id=%s" % sid)
            audit_log(cur, staff, 'delete_contract', 'saving', sid, sr[0], '', ip)
            conn.commit()
            return {'success': True}

        elif action == 'delete_all_transactions':
            sid = int(body['saving_id'])
            cur.execute("SELECT amount FROM savings WHERE id=%s" % sid)
            sr = cur.fetchone()
            if not sr:
                return {'error': 'Договор не найден'}
            cur.execute("DELETE FROM savings_transactions WHERE saving_id=%s" % sid)
            cur.execute("DELETE FROM savings_rate_changes WHERE saving_id=%s" % sid)
            cur.execute("UPDATE savings_schedule SET status='pending', paid_date=NULL, paid_amount=0 WHERE saving_id=%s" % sid)
            orig = float(sr[0])
            cur.execute("UPDATE savings SET current_balance=%s, accrued_interest=0, paid_interest=0, status='active', updated_at=NOW() WHERE id=%s" % (orig, sid))
            audit_log(cur, staff, 'delete_all_transactions', 'saving', sid, '', '', ip)
            conn.commit()
            return {'success': True}

        elif action == 'early_close':
            sid = int(body['saving_id'])
            cur.execute("SELECT amount, accrued_interest, paid_interest, current_balance FROM savings WHERE id=%s" % sid)
            sv = cur.fetchone()
            oa, paid, bal = Decimal(str(sv[0])), Decimal(str(sv[2])), Decimal(str(sv[3]))
            ei = (oa * Decimal('0.001')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            overpaid = paid - ei
            fa = bal - overpaid if overpaid > 0 else bal
            cur.execute("UPDATE savings SET status='early_closed', current_balance=%s, accrued_interest=%s, updated_at=NOW() WHERE id=%s" % (float(fa), float(ei), sid))
            cur.execute("INSERT INTO savings_transactions (saving_id,transaction_date,amount,transaction_type,description) VALUES (%s,'%s',%s,'early_close','Досрочное закрытие')" % (sid, date.today().isoformat(), float(fa)))
            audit_log(cur, staff, 'early_close', 'saving', sid, '', 'Возврат: %s' % float(fa), ip)
            conn.commit()
            return {'success': True, 'final_amount': float(fa), 'early_interest': float(ei)}

def handle_shares(method, params, body, cur, conn, staff=None, ip=''):
    if method == 'GET':
        action = params.get('action', 'list')
        if action == 'detail':
            acc = query_one(cur, "SELECT * FROM share_accounts WHERE id=%s" % params['id'])
            if not acc: return None
            cur.execute("SELECT CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name) ELSE m.company_name END FROM members m WHERE m.id=%s" % acc['member_id'])
            nr = cur.fetchone()
            acc['member_name'] = nr[0] if nr else ''
            if acc.get('org_id'):
                org_row = query_one(cur, "SELECT name, short_name FROM organizations WHERE id=%s" % acc['org_id'])
                acc['org_name'] = org_row['name'] if org_row else ''
                acc['org_short_name'] = org_row['short_name'] if org_row else ''
            acc['transactions'] = query_rows(cur, "SELECT * FROM share_transactions WHERE account_id=%s ORDER BY transaction_date DESC" % params['id'])
            return acc
        else:
            return query_rows(cur, """
                SELECT sa.id, sa.account_no, sa.balance, sa.total_in, sa.total_out, sa.status, sa.created_at, sa.updated_at,
                       sa.org_id,
                       CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name)
                            ELSE m.company_name END as member_name, m.id as member_id,
                       o.name as org_name, o.short_name as org_short_name
                FROM share_accounts sa JOIN members m ON m.id=sa.member_id
                LEFT JOIN organizations o ON o.id=sa.org_id
                ORDER BY sa.created_at DESC
            """)

    elif method == 'POST':
        action = body.get('action', 'create')
        if action == 'create':
            mid = int(body['member_id'])
            a = safe_float(body.get('amount', 0), 'сумма')
            org_id = body.get('org_id')
            cur.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM share_accounts")
            ni = cur.fetchone()[0]
            ano = 'ПС-%06d' % ni
            cur.execute("INSERT INTO share_accounts (account_no,member_id,balance,total_in,org_id) VALUES ('%s',%s,%s,%s,%s) RETURNING id, account_no" % (ano, mid, a, a, org_id if org_id else 'NULL'))
            result = cur.fetchone()
            if a > 0:
                cur.execute("INSERT INTO share_transactions (account_id,transaction_date,amount,transaction_type,description) VALUES (%s,'%s',%s,'in','Первоначальный паевой взнос')" % (result[0], date.today().isoformat(), a))
            audit_log(cur, staff, 'create', 'share', result[0], ano, 'Сумма: %s' % a, ip)
            conn.commit()
            return {'id': result[0], 'account_no': result[1]}
        elif action == 'transaction':
            aid = int(body['account_id'])
            a = safe_float(body['amount'], 'сумма')
            tt = body['transaction_type']
            td = body.get('transaction_date', date.today().isoformat())
            d = body.get('description', '')
            if tt == 'in':
                cur.execute("UPDATE share_accounts SET balance=balance+%s, total_in=total_in+%s, updated_at=NOW() WHERE id=%s" % (a, a, aid))
            else:
                cur.execute("SELECT balance FROM share_accounts WHERE id=%s" % aid)
                if float(cur.fetchone()[0]) < a:
                    return {'error': 'Недостаточно средств'}
                cur.execute("UPDATE share_accounts SET balance=balance-%s, total_out=total_out+%s, updated_at=NOW() WHERE id=%s" % (a, a, aid))
            cur.execute("INSERT INTO share_transactions (account_id,transaction_date,amount,transaction_type,description) VALUES (%s,'%s',%s,'%s','%s')" % (aid, td, a, tt, esc(d)))
            tt_label = 'Внесение' if tt == 'in' else 'Выплата'
            audit_log(cur, staff, 'transaction', 'share', aid, '', '%s: %s' % (tt_label, a), ip)
            conn.commit()
            return {'success': True}

        elif action == 'update_transaction':
            tid = int(body['transaction_id'])
            cur.execute("SELECT account_id, amount, transaction_type FROM share_transactions WHERE id=%s" % tid)
            old = cur.fetchone()
            if not old:
                return {'error': 'Операция не найдена'}
            aid, old_amount, old_tt = old[0], Decimal(str(old[1])), old[2]
            new_amount = Decimal(str(body.get('amount', float(old_amount))))
            new_date = body.get('transaction_date')
            new_desc = body.get('description', '')
            delta = new_amount - old_amount
            upd = "amount=%s" % float(new_amount)
            if new_date:
                upd += ", transaction_date='%s'" % new_date
            if new_desc is not None:
                upd += ", description='%s'" % esc(new_desc)
            cur.execute("UPDATE share_transactions SET %s WHERE id=%s" % (upd, tid))
            if delta != 0:
                if old_tt == 'in':
                    cur.execute("UPDATE share_accounts SET balance=balance+%s, total_in=total_in+%s, updated_at=NOW() WHERE id=%s" % (float(delta), float(delta), aid))
                else:
                    cur.execute("UPDATE share_accounts SET balance=balance-%s, total_out=total_out+%s, updated_at=NOW() WHERE id=%s" % (float(delta), float(delta), aid))
            audit_log(cur, staff, 'update_transaction', 'share', aid, '', 'Транзакция #%s: %s -> %s' % (tid, float(old_amount), float(new_amount)), ip)
            conn.commit()
            return {'success': True}

        elif action == 'delete_transaction':
            tid = int(body['transaction_id'])
            cur.execute("SELECT account_id, amount, transaction_type FROM share_transactions WHERE id=%s" % tid)
            old = cur.fetchone()
            if not old:
                return {'error': 'Операция не найдена'}
            aid, old_amount, old_tt = old[0], Decimal(str(old[1])), old[2]
            cur.execute("DELETE FROM share_transactions WHERE id=%s" % tid)
            if old_tt == 'in':
                cur.execute("UPDATE share_accounts SET balance=balance-%s, total_in=total_in-%s, updated_at=NOW() WHERE id=%s" % (float(old_amount), float(old_amount), aid))
            else:
                cur.execute("UPDATE share_accounts SET balance=balance+%s, total_out=total_out-%s, updated_at=NOW() WHERE id=%s" % (float(old_amount), float(old_amount), aid))
            audit_log(cur, staff, 'delete_transaction', 'share', aid, '', 'Удалена: %s %s' % (old_tt, float(old_amount)), ip)
            conn.commit()
            return {'success': True}

        elif action == 'delete_account':
            aid = int(body['account_id'])
            cur.execute("SELECT account_no FROM share_accounts WHERE id=%s" % aid)
            ar = cur.fetchone()
            if not ar:
                return {'error': 'Счёт не найден'}
            cur.execute("DELETE FROM share_transactions WHERE account_id=%s" % aid)
            cur.execute("DELETE FROM share_accounts WHERE id=%s" % aid)
            audit_log(cur, staff, 'delete_account', 'share', aid, ar[0], '', ip)
            conn.commit()
            return {'success': True}

        elif action == 'delete_all_transactions':
            aid = int(body['account_id'])
            cur.execute("SELECT id FROM share_accounts WHERE id=%s" % aid)
            if not cur.fetchone():
                return {'error': 'Счёт не найден'}
            cur.execute("DELETE FROM share_transactions WHERE account_id=%s" % aid)
            cur.execute("UPDATE share_accounts SET balance=0, total_in=0, total_out=0, updated_at=NOW() WHERE id=%s" % aid)
            audit_log(cur, staff, 'delete_all_transactions', 'share', aid, '', '', ip)
            conn.commit()
            return {'success': True}

def fmt_date(d):
    if not d:
        return ''
    if isinstance(d, str):
        parts = d.split('-')
        if len(parts) == 3:
            return '%s.%s.%s' % (parts[2], parts[1], parts[0])
        return d
    return d.strftime('%d.%m.%Y')

def fmt_money(n):
    if n is None:
        return '0.00'
    return '{:,.2f}'.format(float(n)).replace(',', ' ')

_font_registered = False

def register_cyrillic_font():
    global _font_registered
    if _font_registered:
        return 'DejaVuSans', 'DejaVuSans-Bold'
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    paths = [
        ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
        ('/usr/share/fonts/dejavu-sans-fonts/DejaVuSans.ttf', '/usr/share/fonts/dejavu-sans-fonts/DejaVuSans-Bold.ttf'),
    ]
    for reg_path, bold_path in paths:
        if os.path.exists(reg_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', reg_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_path if os.path.exists(bold_path) else reg_path))
            _font_registered = True
            return 'DejaVuSans', 'DejaVuSans-Bold'
    import urllib.request
    for name, fn in [('DejaVuSans', 'DejaVuSans.ttf'), ('DejaVuSans-Bold', 'DejaVuSans-Bold.ttf')]:
        tmp = '/tmp/%s' % fn
        if not os.path.exists(tmp):
            urllib.request.urlretrieve('https://raw.githubusercontent.com/prawnpdf/prawn/master/data/fonts/%s' % fn, tmp)
        pdfmetrics.registerFont(TTFont(name, tmp))
    _font_registered = True
    return 'DejaVuSans', 'DejaVuSans-Bold'

_logo_cache = {}
_DEFAULT_LOGO_URL = 'https://cdn.poehali.dev/projects/e404b5e6-12a9-4922-a20d-e3c26e46e7a6/bucket/39b830d8-2ba0-408a-8ced-fe6a9eaf99e4.jpg'

def get_logo_path(logo_url=None):
    import urllib.request
    url = logo_url or _DEFAULT_LOGO_URL
    if url in _logo_cache and os.path.exists(_logo_cache[url]):
        return _logo_cache[url]
    ext = url.rsplit('.', 1)[-1] if '.' in url.split('/')[-1] else 'jpg'
    h = hashlib.md5(url.encode()).hexdigest()[:10]
    p = '/tmp/logo_%s.%s' % (h, ext)
    if not os.path.exists(p):
        urllib.request.urlretrieve(url, p)
    _logo_cache[url] = p
    return p

def load_org_settings(cur):
    cur.execute("SELECT key, value FROM organization_settings ORDER BY id")
    rows = cur.fetchall()
    return {r[0]: r[1] for r in rows}

def build_pdf_header(font_r, font_b, org=None):
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT

    if org is None:
        org = {}
    org_name = org.get('name') or 'КПК «ЭКСПЕРТ ФИНАНС»'
    org_phone = org.get('phone') or '8 (800) 700-89-09'
    contacts = []
    if org.get('website'):
        contacts.append('Сайт: %s' % org['website'])
    if org.get('email'):
        contacts.append('Email: %s' % org['email'])
    if org.get('telegram'):
        contacts.append('Telegram: %s' % org['telegram'])
    if org.get('whatsapp'):
        contacts.append('WhatsApp: %s' % org['whatsapp'])
    if not contacts:
        contacts = ['Сайт: nfofinans.ru', 'Email: info@sll-expert.ru', 'Telegram: @nfofinans_161', 'WhatsApp: +79613032756']
    contacts_line = '    '.join(contacts)

    logo_url = org.get('logo_url') or None
    logo_path = get_logo_path(logo_url)
    logo = Image(logo_path, width=24*mm, height=24*mm)

    name_s = ParagraphStyle('HN', fontName=font_b, fontSize=12, leading=14, textColor=colors.HexColor('#1a3c5e'))
    slogan_s = ParagraphStyle('HS', fontName=font_r, fontSize=7, leading=9, textColor=colors.HexColor('#888888'), spaceAfter=1)
    phone_s = ParagraphStyle('HP', fontName=font_b, fontSize=9, leading=11, textColor=colors.HexColor('#333333'), spaceAfter=1)
    contact_s = ParagraphStyle('HC', fontName=font_r, fontSize=6.5, leading=9, textColor=colors.HexColor('#555555'))

    inner_data = [
        [Paragraph(org_name, name_s)],
        [Paragraph('Работаем с финансами, думаем о людях', slogan_s)],
        [Paragraph('Тел: %s' % org_phone, phone_s)],
        [Paragraph(contacts_line, contact_s)],
    ]
    inner = Table(inner_data, colWidths=[None])
    inner.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0), ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))

    header = Table([[logo, inner]], colWidths=[28*mm, None])
    header.setStyle(TableStyle([
        ('VALIGN', (0, 0), (0, 0), 'TOP'), ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0), ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    line_data = [['']]
    line = Table(line_data, colWidths=[186*mm])
    line.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 0.8, colors.HexColor('#2e5d8a')),
        ('TOPPADDING', (0, 0), (-1, -1), 0), ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    return [header, Spacer(1, 3), line, Spacer(1, 8)]

def build_xlsx_header(ws, org=None):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.drawing.image import Image as XlImage

    if org is None:
        org = {}
    org_name = org.get('name') or 'КПК «ЭКСПЕРТ ФИНАНС»'
    org_phone = org.get('phone') or '8 (800) 700-89-09'
    contacts = []
    if org.get('website'):
        contacts.append('Сайт: %s' % org['website'])
    if org.get('email'):
        contacts.append('Email: %s' % org['email'])
    if org.get('telegram'):
        contacts.append('Telegram: %s' % org['telegram'])
    if org.get('whatsapp'):
        contacts.append('WhatsApp: %s' % org['whatsapp'])
    if not contacts:
        contacts = ['Сайт: nfofinans.ru', 'Email: info@sll-expert.ru', 'Telegram: @nfofinans_161', 'WhatsApp: +79613032756']
    contacts_line = '    '.join(contacts)

    logo_url = org.get('logo_url') or None
    logo_path = get_logo_path(logo_url)
    img = XlImage(logo_path)
    img.width = 80
    img.height = 80
    ws.add_image(img, 'A1')
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 12
    ws.merge_cells('B1:F2')
    ws['B1'] = org_name
    ws['B1'].font = Font(bold=True, size=14, color='1a3c5e')
    ws['B1'].alignment = Alignment(vertical='center')
    ws['B3'] = 'Работаем с финансами, думаем о людях'
    ws['B3'].font = Font(italic=True, size=8, color='888888')
    ws['B4'] = 'Тел: %s' % org_phone
    ws['B4'].font = Font(bold=True, size=9, color='333333')
    ws['B5'] = contacts_line
    ws['B5'].font = Font(size=7, color='555555')
    line_border = Border(bottom=Side(style='medium', color='2e5d8a'))
    for col in range(1, 8):
        ws.cell(row=6, column=col).border = line_border
    ws.row_dimensions[6].height = 5
    return 8

def generate_loan_xlsx(loan, schedule, payments, member_name, org=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Выписка по займу'

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    row = build_xlsx_header(ws, org)

    ws.merge_cells('A%d:F%d' % (row, row))
    ws['A%d' % row] = 'Выписка по договору займа %s' % loan.get('contract_no', '')
    ws['A%d' % row].font = title_font
    row += 2

    ws['A%d' % row] = 'Пайщик:'
    ws['B%d' % row] = member_name
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Сумма займа:'
    ws['B%d' % row] = '%s руб.' % fmt_money(loan.get('amount'))
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Ставка:'
    ws['B%d' % row] = '%s%% годовых' % loan.get('rate', '')
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Срок:'
    ws['B%d' % row] = '%s мес.' % loan.get('term_months', '')
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Период:'
    ws['B%d' % row] = '%s — %s' % (fmt_date(loan.get('start_date')), fmt_date(loan.get('end_date')))
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Остаток:'
    ws['B%d' % row] = '%s руб.' % fmt_money(loan.get('balance'))
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Статус:'
    status_map = {'active': 'Активен', 'closed': 'Закрыт', 'overdue': 'Просрочен'}
    ws['B%d' % row] = status_map.get(loan.get('status', ''), loan.get('status', ''))
    ws['A%d' % row].font = Font(bold=True)
    row += 2

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14

    ws.merge_cells('A%d:G%d' % (row, row))
    ws['A%d' % row] = 'ГРАФИК ПЛАТЕЖЕЙ'
    ws['A%d' % row].font = Font(bold=True, size=12)
    row += 1

    sched_headers = ['№', 'Дата', 'Платёж', 'Осн. долг', 'Проценты', 'Остаток', 'Статус']
    for ci, h in enumerate(sched_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
    row += 1

    status_labels = {'pending': 'Ожидается', 'paid': 'Оплачен', 'partial': 'Частично', 'overdue': 'Просрочен'}
    for item in schedule:
        ws.cell(row=row, column=1, value=item.get('payment_no')).border = border
        ws.cell(row=row, column=2, value=fmt_date(item.get('payment_date'))).border = border
        ws.cell(row=row, column=3, value=float(item.get('payment_amount', 0))).border = border
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=float(item.get('principal_amount', 0))).border = border
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=float(item.get('interest_amount', 0))).border = border
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=float(item.get('balance_after', 0))).border = border
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=status_labels.get(item.get('status', 'pending'), item.get('status', ''))).border = border
        row += 1

    total_payment = sum(float(i.get('payment_amount', 0)) for i in schedule)
    total_principal = sum(float(i.get('principal_amount', 0)) for i in schedule)
    total_interest = sum(float(i.get('interest_amount', 0)) for i in schedule)
    ws.cell(row=row, column=1, value='ИТОГО').font = Font(bold=True)
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2).border = border
    ws.cell(row=row, column=3, value=total_payment).border = border
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_principal).border = border
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    ws.cell(row=row, column=5, value=total_interest).border = border
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    ws.cell(row=row, column=6).border = border
    ws.cell(row=row, column=7).border = border
    row += 2

    if payments:
        ws.merge_cells('A%d:F%d' % (row, row))
        ws['A%d' % row] = 'ИСТОРИЯ ПЛАТЕЖЕЙ'
        ws['A%d' % row].font = Font(bold=True, size=12)
        row += 1

        pay_headers = ['Дата', 'Сумма', 'Осн. долг', 'Проценты', 'Штрафы', 'Тип']
        for ci, h in enumerate(pay_headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = Alignment(horizontal='center')
        row += 1

        for p in payments:
            ws.cell(row=row, column=1, value=fmt_date(p.get('payment_date'))).border = border
            ws.cell(row=row, column=2, value=float(p.get('amount', 0))).border = border
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=float(p.get('principal_part', 0))).border = border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=float(p.get('interest_part', 0))).border = border
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=float(p.get('penalty_part', 0))).border = border
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            type_labels = {'regular': 'Обычный', 'early_full': 'Досрочное полное', 'early_partial': 'Досрочное частичное'}
            ws.cell(row=row, column=6, value=type_labels.get(p.get('payment_type', ''), p.get('payment_type', ''))).border = border
            row += 1

    row += 1
    ws['A%d' % row] = 'Дата формирования: %s' % datetime.now().strftime('%d.%m.%Y %H:%M')
    ws['A%d' % row].font = Font(italic=True, color='666666')

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def generate_loan_pdf(loan, schedule, payments, member_name, org=None):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    font_r, font_b = register_cyrillic_font()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story = []

    story.extend(build_pdf_header(font_r, font_b, org))

    title_style = ParagraphStyle('T', fontName=font_b, fontSize=13, spaceAfter=4, textColor=colors.HexColor('#1a3c5e'))
    sub_style = ParagraphStyle('S', fontName=font_b, fontSize=10, spaceAfter=4, spaceBefore=8, textColor=colors.HexColor('#2e5d8a'))
    footer_style = ParagraphStyle('F', fontName=font_r, fontSize=7, textColor=colors.grey)

    story.append(Paragraph('Выписка по договору займа %s' % loan.get('contract_no', ''), title_style))
    story.append(Spacer(1, 4))

    status_map = {'active': 'Активен', 'closed': 'Закрыт', 'overdue': 'Просрочен'}
    info_data = [
        ['Пайщик:', member_name, 'Сумма:', '%s руб.' % fmt_money(loan.get('amount'))],
        ['Ставка:', '%s%% годовых' % loan.get('rate', ''), 'Срок:', '%s мес.' % loan.get('term_months', '')],
        ['Период:', '%s — %s' % (fmt_date(loan.get('start_date')), fmt_date(loan.get('end_date'))), 'Остаток:', '%s руб.' % fmt_money(loan.get('balance'))],
        ['Статус:', status_map.get(loan.get('status', ''), loan.get('status', '')), '', ''],
    ]
    info_table = Table(info_data, colWidths=[55, 150, 55, 150])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_r), ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (0, -1), font_b), ('FONTNAME', (2, 0), (2, -1), font_b),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2), ('TOPPADDING', (0, 0), (-1, -1), 1),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 6))

    story.append(Paragraph('График платежей', sub_style))
    sched_data = [['№', 'Дата', 'Платёж', 'Осн. долг', 'Проценты', 'Остаток', 'Статус']]
    status_labels = {'pending': 'Ожидается', 'paid': 'Оплачен', 'partial': 'Частично', 'overdue': 'Просрочен'}
    for item in schedule:
        sched_data.append([
            str(item.get('payment_no', '')), fmt_date(item.get('payment_date')),
            fmt_money(item.get('payment_amount', 0)), fmt_money(item.get('principal_amount', 0)),
            fmt_money(item.get('interest_amount', 0)), fmt_money(item.get('balance_after', 0)),
            status_labels.get(item.get('status', 'pending'), item.get('status', '')),
        ])
    total_payment = sum(float(i.get('payment_amount', 0)) for i in schedule)
    total_principal = sum(float(i.get('principal_amount', 0)) for i in schedule)
    total_interest = sum(float(i.get('interest_amount', 0)) for i in schedule)
    sched_data.append(['ИТОГО', '', fmt_money(total_payment), fmt_money(total_principal), fmt_money(total_interest), '', ''])

    cw = [22, 58, 68, 68, 60, 68, 58]
    st = Table(sched_data, colWidths=cw, repeatRows=1)
    st.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_r), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('FONTNAME', (0, 0), (-1, 0), font_b), ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dce6f0')),
        ('FONTNAME', (0, -1), (-1, -1), font_b), ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#eef2f7')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#b0b0b0')),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'), ('ALIGN', (2, 0), (5, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fb')]),
    ]))
    story.append(st)

    if payments:
        story.append(Spacer(1, 6))
        story.append(Paragraph('История платежей', sub_style))
        pay_data = [['Дата', 'Сумма', 'Осн. долг', 'Проценты', 'Штрафы', 'Тип']]
        type_labels = {'regular': 'Обычный', 'early_full': 'Досрочное полное', 'early_partial': 'Досрочное частичное'}
        for p in payments:
            pay_data.append([
                fmt_date(p.get('payment_date')), fmt_money(p.get('amount', 0)),
                fmt_money(p.get('principal_part', 0)), fmt_money(p.get('interest_part', 0)),
                fmt_money(p.get('penalty_part', 0)),
                type_labels.get(p.get('payment_type', ''), p.get('payment_type', '')),
            ])
        pt = Table(pay_data, colWidths=[58, 68, 68, 60, 60, 100], repeatRows=1)
        pt.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_r), ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('FONTNAME', (0, 0), (-1, 0), font_b), ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dce6f0')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#b0b0b0')),
            ('ALIGN', (1, 0), (4, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fb')]),
        ]))
        story.append(pt)

    story.append(Spacer(1, 12))
    story.append(Paragraph('Дата формирования: %s' % datetime.now().strftime('%d.%m.%Y %H:%M'), footer_style))

    doc.build(story)
    return buf.getvalue()

def generate_savings_xlsx(saving, schedule, transactions, member_name, org=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Выписка по сбережению'

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')

    row = build_xlsx_header(ws, org)

    ws.merge_cells('A%d:F%d' % (row, row))
    ws['A%d' % row] = 'Выписка по договору сбережений %s' % saving.get('contract_no', '')
    ws['A%d' % row].font = Font(bold=True, size=14)
    row += 2

    ws['A%d' % row] = 'Пайщик:'
    ws['B%d' % row] = member_name
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Сумма вклада:'
    ws['B%d' % row] = '%s руб.' % fmt_money(saving.get('amount'))
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Ставка:'
    ws['B%d' % row] = '%s%% годовых' % saving.get('rate', '')
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Срок:'
    ws['B%d' % row] = '%s мес.' % saving.get('term_months', '')
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Период:'
    ws['B%d' % row] = '%s — %s' % (fmt_date(saving.get('start_date')), fmt_date(saving.get('end_date')))
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Начислено %:'
    ws['B%d' % row] = '%s руб.' % fmt_money(saving.get('accrued_interest'))
    ws['A%d' % row].font = Font(bold=True)
    row += 2

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    ws['A%d' % row] = 'ГРАФИК ДОХОДНОСТИ'
    ws['A%d' % row].font = Font(bold=True, size=12)
    row += 1

    headers = ['№', 'Начало', 'Окончание', 'Проценты', 'Накоплено', 'Баланс']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
    row += 1

    for item in schedule:
        ws.cell(row=row, column=1, value=item.get('period_no')).border = border
        ws.cell(row=row, column=2, value=fmt_date(item.get('period_start'))).border = border
        ws.cell(row=row, column=3, value=fmt_date(item.get('period_end'))).border = border
        ws.cell(row=row, column=4, value=float(item.get('interest_amount', 0))).border = border
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=float(item.get('cumulative_interest', 0))).border = border
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=float(item.get('balance_after', 0))).border = border
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        row += 1

    if transactions:
        row += 1
        ws['A%d' % row] = 'ОПЕРАЦИИ'
        ws['A%d' % row].font = Font(bold=True, size=12)
        row += 1
        t_headers = ['Дата', 'Сумма', 'Тип', 'Описание']
        for ci, h in enumerate(t_headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
        row += 1
        type_labels = {'deposit': 'Пополнение', 'withdrawal': 'Снятие', 'interest_payout': 'Выплата %', 'early_close': 'Досрочное закрытие'}
        for t in transactions:
            ws.cell(row=row, column=1, value=fmt_date(t.get('transaction_date'))).border = border
            ws.cell(row=row, column=2, value=float(t.get('amount', 0))).border = border
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=type_labels.get(t.get('transaction_type', ''), t.get('transaction_type', ''))).border = border
            ws.cell(row=row, column=4, value=t.get('description', '')).border = border
            row += 1

    row += 1
    ws['A%d' % row] = 'Дата формирования: %s' % datetime.now().strftime('%d.%m.%Y %H:%M')
    ws['A%d' % row].font = Font(italic=True, color='666666')

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def generate_savings_pdf(saving, schedule, transactions, member_name, org=None):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    font_r, font_b = register_cyrillic_font()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story = []

    story.extend(build_pdf_header(font_r, font_b, org))

    title_style = ParagraphStyle('T', fontName=font_b, fontSize=13, spaceAfter=4, textColor=colors.HexColor('#1a3c5e'))
    sub_style = ParagraphStyle('S', fontName=font_b, fontSize=10, spaceAfter=4, spaceBefore=8, textColor=colors.HexColor('#2e5d8a'))
    footer_style = ParagraphStyle('F', fontName=font_r, fontSize=7, textColor=colors.grey)

    story.append(Paragraph('Выписка по договору сбережений %s' % saving.get('contract_no', ''), title_style))
    story.append(Spacer(1, 4))

    info = [
        ['Пайщик:', member_name, 'Сумма:', '%s руб.' % fmt_money(saving.get('amount'))],
        ['Ставка:', '%s%% годовых' % saving.get('rate', ''), 'Срок:', '%s мес.' % saving.get('term_months', '')],
        ['Период:', '%s — %s' % (fmt_date(saving.get('start_date')), fmt_date(saving.get('end_date'))), 'Начислено:', '%s руб.' % fmt_money(saving.get('accrued_interest'))],
    ]
    it = Table(info, colWidths=[55, 150, 60, 150])
    it.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_r), ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (0, -1), font_b), ('FONTNAME', (2, 0), (2, -1), font_b),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2), ('TOPPADDING', (0, 0), (-1, -1), 1),
    ]))
    story.append(it)
    story.append(Spacer(1, 6))

    story.append(Paragraph('График доходности', sub_style))
    sdata = [['№', 'Начало', 'Окончание', 'Проценты', 'Накоплено', 'Баланс']]
    for item in schedule:
        sdata.append([str(item.get('period_no', '')), fmt_date(item.get('period_start')), fmt_date(item.get('period_end')),
                       fmt_money(item.get('interest_amount', 0)), fmt_money(item.get('cumulative_interest', 0)), fmt_money(item.get('balance_after', 0))])
    st = Table(sdata, colWidths=[22, 68, 68, 75, 75, 85], repeatRows=1)
    st.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_r), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('FONTNAME', (0, 0), (-1, 0), font_b), ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d6eaf8')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#b0b0b0')),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'), ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f6fc')]),
    ]))
    story.append(st)

    if transactions:
        story.append(Spacer(1, 6))
        story.append(Paragraph('Операции', sub_style))
        tdata = [['Дата', 'Сумма', 'Тип', 'Описание']]
        type_labels = {'deposit': 'Пополнение', 'withdrawal': 'Снятие', 'interest_payout': 'Выплата %', 'early_close': 'Досрочное закрытие'}
        for t in transactions:
            tdata.append([fmt_date(t.get('transaction_date')), fmt_money(t.get('amount', 0)),
                           type_labels.get(t.get('transaction_type', ''), t.get('transaction_type', '')), t.get('description', '')])
        tt = Table(tdata, colWidths=[58, 68, 100, 170], repeatRows=1)
        tt.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_r), ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('FONTNAME', (0, 0), (-1, 0), font_b), ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d6eaf8')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#b0b0b0')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f6fc')]),
        ]))
        story.append(tt)

    story.append(Spacer(1, 12))
    story.append(Paragraph('Дата формирования: %s' % datetime.now().strftime('%d.%m.%Y %H:%M'), footer_style))
    doc.build(story)
    return buf.getvalue()

def generate_saving_transactions_xlsx(saving, transactions, member_name, org=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Выписка по транзакциям'

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')

    row = build_xlsx_header(ws, org)

    ws.merge_cells('A%d:F%d' % (row, row))
    ws['A%d' % row] = 'Выписка по транзакциям — договор %s' % saving.get('contract_no', '')
    ws['A%d' % row].font = Font(bold=True, size=14)
    row += 2

    ws['A%d' % row] = 'Пайщик:'
    ws['B%d' % row] = member_name
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Сумма вклада:'
    ws['B%d' % row] = '%s руб.' % fmt_money(saving.get('amount'))
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Ставка:'
    ws['B%d' % row] = '%s%% годовых' % saving.get('rate', '')
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Срок:'
    ws['B%d' % row] = '%s мес.' % saving.get('term_months', '')
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Период:'
    ws['B%d' % row] = '%s — %s' % (fmt_date(saving.get('start_date')), fmt_date(saving.get('end_date')))
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Текущий баланс:'
    ws['B%d' % row] = '%s руб.' % fmt_money(saving.get('current_balance'))
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    status_map = {'active': 'Активен', 'closed': 'Закрыт', 'early_closed': 'Досрочно закрыт'}
    ws['A%d' % row] = 'Статус:'
    ws['B%d' % row] = status_map.get(saving.get('status', ''), saving.get('status', ''))
    ws['A%d' % row].font = Font(bold=True)
    row += 2

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 40

    ws['A%d' % row] = 'ТРАНЗАКЦИИ'
    ws['A%d' % row].font = Font(bold=True, size=12)
    row += 1

    type_labels = {'opening': 'Открытие', 'deposit': 'Пополнение', 'withdrawal': 'Частичное изъятие', 'interest_payout': 'Выплата процентов', 'interest_accrual': 'Начисление процентов', 'term_change': 'Изменение срока', 'rate_change': 'Изменение ставки', 'early_close': 'Досрочное закрытие', 'closing': 'Закрытие'}
    headers = ['№', 'Дата', 'Сумма', 'Тип операции', 'Описание']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
    row += 1

    running_balance = Decimal('0')
    for idx, t in enumerate(transactions, 1):
        tt = t.get('transaction_type', '')
        amt = float(t.get('amount', 0))
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=fmt_date(t.get('transaction_date'))).border = border
        ws.cell(row=row, column=3, value=amt).border = border
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=type_labels.get(tt, tt)).border = border
        ws.cell(row=row, column=5, value=t.get('description', '')).border = border
        row += 1

    row += 1
    ws['A%d' % row] = 'Всего транзакций: %d' % len(transactions)
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Дата формирования: %s' % datetime.now().strftime('%d.%m.%Y %H:%M')
    ws['A%d' % row].font = Font(italic=True, color='666666')

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def generate_saving_transactions_pdf(saving, transactions, member_name, org=None):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    font_r, font_b = register_cyrillic_font()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=15*mm, bottomMargin=15*mm)
    story = []

    story.extend(build_pdf_header(font_r, font_b, org))

    title_style = ParagraphStyle('T', fontName=font_b, fontSize=13, spaceAfter=4, textColor=colors.HexColor('#1a3c5e'))
    sub_style = ParagraphStyle('S', fontName=font_b, fontSize=10, spaceAfter=4, spaceBefore=8, textColor=colors.HexColor('#2e5d8a'))
    footer_style = ParagraphStyle('F', fontName=font_r, fontSize=7, textColor=colors.grey)
    desc_style = ParagraphStyle('D', fontName=font_r, fontSize=6.5, leading=8)

    story.append(Paragraph('Выписка по транзакциям — договор %s' % saving.get('contract_no', ''), title_style))
    story.append(Spacer(1, 4))

    status_map = {'active': 'Активен', 'closed': 'Закрыт', 'early_closed': 'Досрочно закрыт'}
    info = [
        ['Пайщик:', member_name, 'Сумма:', '%s руб.' % fmt_money(saving.get('amount'))],
        ['Ставка:', '%s%% годовых' % saving.get('rate', ''), 'Срок:', '%s мес.' % saving.get('term_months', '')],
        ['Период:', '%s — %s' % (fmt_date(saving.get('start_date')), fmt_date(saving.get('end_date'))), 'Баланс:', '%s руб.' % fmt_money(saving.get('current_balance'))],
        ['Статус:', status_map.get(saving.get('status', ''), saving.get('status', '')), '', ''],
    ]
    it = Table(info, colWidths=[55, 150, 55, 150])
    it.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_r), ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (0, -1), font_b), ('FONTNAME', (2, 0), (2, -1), font_b),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2), ('TOPPADDING', (0, 0), (-1, -1), 1),
    ]))
    story.append(it)
    story.append(Spacer(1, 6))

    story.append(Paragraph('Транзакции', sub_style))
    type_labels = {'opening': 'Открытие', 'deposit': 'Пополнение', 'withdrawal': 'Частичное изъятие', 'interest_payout': 'Выплата %', 'interest_accrual': 'Начисление %', 'term_change': 'Изм. срока', 'rate_change': 'Изм. ставки', 'early_close': 'Досрочное закр.', 'closing': 'Закрытие'}
    tdata = [['№', 'Дата', 'Сумма', 'Тип', 'Описание']]
    for idx, t in enumerate(transactions, 1):
        tt = t.get('transaction_type', '')
        desc_text = t.get('description', '') or ''
        tdata.append([str(idx), fmt_date(t.get('transaction_date')), fmt_money(t.get('amount', 0)),
                       type_labels.get(tt, tt), Paragraph(desc_text, desc_style)])
    tt_table = Table(tdata, colWidths=[22, 58, 68, 75, 170], repeatRows=1)
    tt_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_r), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('FONTNAME', (0, 0), (-1, 0), font_b), ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8daef')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#b0b0b0')),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'), ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f0fc')]),
    ]))
    story.append(tt_table)

    story.append(Spacer(1, 8))
    story.append(Paragraph('Всего транзакций: %d' % len(transactions), ParagraphStyle('C', fontName=font_b, fontSize=8)))
    story.append(Spacer(1, 12))
    story.append(Paragraph('Дата формирования: %s' % datetime.now().strftime('%d.%m.%Y %H:%M'), footer_style))
    doc.build(story)
    return buf.getvalue()

def generate_shares_xlsx(account, transactions, member_name, org=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Выписка по паевому счёту'

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    row = build_xlsx_header(ws, org)

    ws.merge_cells('A%d:D%d' % (row, row))
    ws['A%d' % row] = 'Выписка по паевому счёту %s' % account.get('account_no', '')
    ws['A%d' % row].font = Font(bold=True, size=14)
    row += 2

    ws['A%d' % row] = 'Пайщик:'
    ws['B%d' % row] = member_name
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Баланс:'
    ws['B%d' % row] = '%s руб.' % fmt_money(account.get('balance'))
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Внесено:'
    ws['B%d' % row] = '%s руб.' % fmt_money(account.get('total_in'))
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws['A%d' % row] = 'Выплачено:'
    ws['B%d' % row] = '%s руб.' % fmt_money(account.get('total_out'))
    ws['A%d' % row].font = Font(bold=True)
    row += 2

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30

    ws['A%d' % row] = 'ОПЕРАЦИИ'
    ws['A%d' % row].font = Font(bold=True, size=12)
    row += 1

    headers = ['Дата', 'Сумма', 'Тип', 'Описание']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
    row += 1

    type_labels = {'in': 'Внесение', 'out': 'Выплата'}
    for t in transactions:
        ws.cell(row=row, column=1, value=fmt_date(t.get('transaction_date'))).border = border
        ws.cell(row=row, column=2, value=float(t.get('amount', 0))).border = border
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=type_labels.get(t.get('transaction_type', ''), t.get('transaction_type', ''))).border = border
        ws.cell(row=row, column=4, value=t.get('description', '')).border = border
        row += 1

    row += 1
    ws['A%d' % row] = 'Дата формирования: %s' % datetime.now().strftime('%d.%m.%Y %H:%M')
    ws['A%d' % row].font = Font(italic=True, color='666666')

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def generate_shares_pdf(account, transactions, member_name, org=None):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    font_r, font_b = register_cyrillic_font()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story = []

    story.extend(build_pdf_header(font_r, font_b, org))

    title_style = ParagraphStyle('T', fontName=font_b, fontSize=13, spaceAfter=4, textColor=colors.HexColor('#1a3c5e'))
    sub_style = ParagraphStyle('S', fontName=font_b, fontSize=10, spaceAfter=4, spaceBefore=8, textColor=colors.HexColor('#2e5d8a'))
    footer_style = ParagraphStyle('F', fontName=font_r, fontSize=7, textColor=colors.grey)

    story.append(Paragraph('Выписка по паевому счёту %s' % account.get('account_no', ''), title_style))
    story.append(Spacer(1, 4))

    info = [['Пайщик:', member_name], ['Баланс:', '%s руб.' % fmt_money(account.get('balance'))],
            ['Внесено:', '%s руб.' % fmt_money(account.get('total_in'))], ['Выплачено:', '%s руб.' % fmt_money(account.get('total_out'))]]
    it = Table(info, colWidths=[70, 200])
    it.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_r), ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (0, -1), font_b),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2), ('TOPPADDING', (0, 0), (-1, -1), 1),
    ]))
    story.append(it)
    story.append(Spacer(1, 6))

    story.append(Paragraph('Операции', sub_style))
    tdata = [['Дата', 'Сумма', 'Тип', 'Описание']]
    type_labels = {'in': 'Внесение', 'out': 'Выплата'}
    for t in transactions:
        tdata.append([fmt_date(t.get('transaction_date')), fmt_money(t.get('amount', 0)),
                       type_labels.get(t.get('transaction_type', ''), t.get('transaction_type', '')), t.get('description', '')])
    tt = Table(tdata, colWidths=[58, 68, 85, 200], repeatRows=1)
    tt.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_r), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('FONTNAME', (0, 0), (-1, 0), font_b), ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fce4d6')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#b0b0b0')),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef5ef')]),
    ]))
    story.append(tt)

    story.append(Spacer(1, 12))
    story.append(Paragraph('Дата формирования: %s' % datetime.now().strftime('%d.%m.%Y %H:%M'), footer_style))
    doc.build(story)
    return buf.getvalue()

def handle_export(params, cur):
    export_type = params.get('type', 'loan')
    format_ = params.get('format', 'xlsx')
    item_id = params.get('id')
    if not item_id:
        return None

    default_org = load_org_settings(cur)

    if export_type == 'loan':
        loan = query_one(cur, "SELECT * FROM loans WHERE id = %s" % item_id)
        if not loan:
            return None
        org_id = loan.get('org_id')
        if org_id:
            org_row = query_one(cur, "SELECT * FROM organizations WHERE id=%s" % org_id)
            org = org_row if org_row else default_org
        else:
            org = default_org
        cur.execute("SELECT CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name) ELSE m.company_name END FROM members m WHERE m.id=%s" % loan['member_id'])
        nr = cur.fetchone()
        member_name = nr[0] if nr else ''
        schedule = query_rows(cur, "SELECT * FROM loan_schedule WHERE loan_id=%s ORDER BY payment_no" % item_id)
        payments = query_rows(cur, "SELECT * FROM loan_payments WHERE loan_id=%s ORDER BY payment_date" % item_id)
        if format_ == 'pdf':
            data = generate_loan_pdf(loan, schedule, payments, member_name, org)
            ct = 'application/pdf'
            fn = 'loan_%s.pdf' % loan.get('contract_no', item_id)
        else:
            data = generate_loan_xlsx(loan, schedule, payments, member_name, org)
            ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            fn = 'loan_%s.xlsx' % loan.get('contract_no', item_id)

    elif export_type == 'saving':
        saving = query_one(cur, "SELECT * FROM savings WHERE id = %s" % item_id)
        if not saving:
            return None
        org_id = saving.get('org_id')
        if org_id:
            org_row = query_one(cur, "SELECT * FROM organizations WHERE id=%s" % org_id)
            org = org_row if org_row else default_org
        else:
            org = default_org
        cur.execute("SELECT CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name) ELSE m.company_name END FROM members m WHERE m.id=%s" % saving['member_id'])
        nr = cur.fetchone()
        member_name = nr[0] if nr else ''
        schedule = query_rows(cur, "SELECT * FROM savings_schedule WHERE saving_id=%s ORDER BY period_no" % item_id)
        transactions = query_rows(cur, "SELECT * FROM savings_transactions WHERE saving_id=%s ORDER BY transaction_date" % item_id)
        if format_ == 'pdf':
            data = generate_savings_pdf(saving, schedule, transactions, member_name, org)
            ct = 'application/pdf'
            fn = 'saving_%s.pdf' % saving.get('contract_no', item_id)
        else:
            data = generate_savings_xlsx(saving, schedule, transactions, member_name, org)
            ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            fn = 'saving_%s.xlsx' % saving.get('contract_no', item_id)

    elif export_type == 'saving_transactions':
        saving = query_one(cur, "SELECT * FROM savings WHERE id = %s" % item_id)
        if not saving:
            return None
        org_id = saving.get('org_id')
        if org_id:
            org_row = query_one(cur, "SELECT * FROM organizations WHERE id=%s" % org_id)
            org = org_row if org_row else default_org
        else:
            org = default_org
        cur.execute("SELECT CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name) ELSE m.company_name END FROM members m WHERE m.id=%s" % saving['member_id'])
        nr = cur.fetchone()
        member_name = nr[0] if nr else ''
        transactions = query_rows(cur, "SELECT * FROM savings_transactions WHERE saving_id=%s ORDER BY transaction_date, id" % item_id)
        if format_ == 'pdf':
            data = generate_saving_transactions_pdf(saving, transactions, member_name, org)
            ct = 'application/pdf'
            fn = 'transactions_%s.pdf' % saving.get('contract_no', item_id)
        else:
            data = generate_saving_transactions_xlsx(saving, transactions, member_name, org)
            ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            fn = 'transactions_%s.xlsx' % saving.get('contract_no', item_id)

    elif export_type == 'share':
        account = query_one(cur, "SELECT * FROM share_accounts WHERE id = %s" % item_id)
        if not account:
            return None
        org_id = account.get('org_id')
        if org_id:
            org_row = query_one(cur, "SELECT * FROM organizations WHERE id=%s" % org_id)
            org = org_row if org_row else default_org
        else:
            org = default_org
        cur.execute("SELECT CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name,' ',m.middle_name) ELSE m.company_name END FROM members m WHERE m.id=%s" % account['member_id'])
        nr = cur.fetchone()
        member_name = nr[0] if nr else ''
        transactions = query_rows(cur, "SELECT * FROM share_transactions WHERE account_id=%s ORDER BY transaction_date DESC" % item_id)
        if format_ == 'pdf':
            data = generate_shares_pdf(account, transactions, member_name, org)
            ct = 'application/pdf'
            fn = 'share_%s.pdf' % account.get('account_no', item_id)
        else:
            data = generate_shares_xlsx(account, transactions, member_name, org)
            ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            fn = 'share_%s.xlsx' % account.get('account_no', item_id)
    else:
        return None

    return {'file': base64.b64encode(data).decode('utf-8'), 'content_type': ct, 'filename': fn}

def handle_dashboard(cur, params=None):
    params = params or {}
    org_id = params.get('org_id')
    org_filter_loans = " AND l.org_id=%s" % org_id if org_id else ""
    org_filter_savings = " AND s.org_id=%s" % org_id if org_id else ""
    org_filter_shares = " AND sa.org_id=%s" % org_id if org_id else ""

    stats = {}

    if org_id:
        cur.execute("SELECT COUNT(DISTINCT m.id) FROM members m JOIN loans l ON l.member_id=m.id WHERE m.status='active' AND l.org_id=%s UNION SELECT COUNT(DISTINCT m.id) FROM members m JOIN savings s ON s.member_id=m.id WHERE m.status='active' AND s.org_id=%s UNION SELECT COUNT(DISTINCT m.id) FROM members m JOIN share_accounts sa ON sa.member_id=m.id WHERE m.status='active' AND sa.org_id=%s" % (org_id, org_id, org_id))
        rows = cur.fetchall()
        member_ids = set()
        cur.execute("SELECT DISTINCT m.id FROM members m LEFT JOIN loans l ON l.member_id=m.id AND l.org_id=%s LEFT JOIN savings s ON s.member_id=m.id AND s.org_id=%s LEFT JOIN share_accounts sa ON sa.member_id=m.id AND sa.org_id=%s WHERE m.status='active' AND (l.id IS NOT NULL OR s.id IS NOT NULL OR sa.id IS NOT NULL)" % (org_id, org_id, org_id))
        stats['total_members'] = len(cur.fetchall())
    else:
        cur.execute("SELECT COUNT(*) FROM members WHERE status='active'")
        stats['total_members'] = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*), COALESCE(SUM(l.balance),0) FROM loans l WHERE l.status='active'%s" % org_filter_loans)
    r = cur.fetchone()
    stats['active_loans'] = r[0]
    stats['loan_portfolio'] = float(r[1])

    cur.execute("SELECT COUNT(*) FROM loans l WHERE l.status='overdue'%s" % org_filter_loans)
    stats['overdue_loans'] = cur.fetchone()[0]

    cur.execute("SELECT COALESCE(SUM(s.current_balance),0) FROM savings s WHERE s.status='active'%s" % org_filter_savings)
    stats['total_savings'] = float(cur.fetchone()[0])

    cur.execute("SELECT COALESCE(SUM(sa.balance),0) FROM share_accounts sa WHERE sa.status='active'%s" % org_filter_shares)
    stats['total_shares'] = float(cur.fetchone()[0])

    cur.execute("SELECT id, name, short_name FROM organizations WHERE is_active=true ORDER BY name")
    stats['organizations'] = [{'id': r[0], 'name': r[1], 'short_name': r[2]} for r in cur.fetchall()]

    overdue_loans = []
    cur.execute("""
        SELECT l.id, l.contract_no, m.id as member_id,
            COALESCE(m.last_name,'') || ' ' || COALESCE(m.first_name,'') || ' ' || COALESCE(m.middle_name,'') as member_name,
            l.balance, l.rate, l.end_date, l.org_id,
            COALESCE(o.short_name, o.name, '') as org_name
        FROM loans l
        JOIN members m ON m.id = l.member_id
        LEFT JOIN organizations o ON o.id = l.org_id
        WHERE l.status = 'overdue'%s
        ORDER BY l.end_date
    """ % org_filter_loans)
    for r in cur.fetchall():
        loan_id = r[0]
        cur.execute("""
            SELECT COALESCE(SUM(ls.payment_amount - COALESCE(ls.paid_amount,0)),0),
                   MIN(ls.payment_date),
                   MAX(CASE WHEN ls.status='overdue' OR (ls.status='pending' AND ls.payment_date < CURRENT_DATE) THEN (CURRENT_DATE - ls.payment_date) ELSE 0 END),
                   COALESCE(SUM(ls.penalty_amount),0)
            FROM loan_schedule ls
            WHERE ls.loan_id=%s AND ls.status IN ('overdue','pending') AND ls.payment_date < CURRENT_DATE
        """ % loan_id)
        sched = cur.fetchone()
        overdue_loans.append({
            'loan_id': r[0], 'contract_no': r[1], 'member_id': r[2], 'member_name': r[3].strip(),
            'balance': float(r[4]), 'rate': float(r[5]), 'end_date': str(r[6]),
            'org_id': r[7], 'org_name': r[8],
            'overdue_amount': float(sched[0]) if sched[0] else 0,
            'overdue_since': str(sched[1]) if sched[1] else None,
            'overdue_days': int(sched[2]) if sched[2] else 0,
            'penalty_total': float(sched[3]) if sched[3] else 0
        })
    stats['overdue_loan_list'] = overdue_loans

    expiring_savings = []
    cur.execute("""
        SELECT s.id, s.contract_no, m.id as member_id,
            COALESCE(m.last_name,'') || ' ' || COALESCE(m.first_name,'') || ' ' || COALESCE(m.middle_name,'') as member_name,
            s.current_balance, s.accrued_interest, s.paid_interest, s.rate, s.end_date,
            s.org_id, COALESCE(o.short_name, o.name, '') as org_name
        FROM savings s
        JOIN members m ON m.id = s.member_id
        LEFT JOIN organizations o ON o.id = s.org_id
        WHERE s.status = 'active'
          AND s.end_date <= CURRENT_DATE + INTERVAL '30 days'
          AND s.end_date >= CURRENT_DATE%s
        ORDER BY s.end_date
    """ % org_filter_savings)
    total_refund = 0
    for r in cur.fetchall():
        refund = float(r[4]) + float(r[5]) - float(r[6])
        total_refund += refund
        expiring_savings.append({
            'saving_id': r[0], 'contract_no': r[1], 'member_id': r[2], 'member_name': r[3].strip(),
            'current_balance': float(r[4]), 'accrued_interest': float(r[5]),
            'paid_interest': float(r[6]), 'rate': float(r[7]), 'end_date': str(r[8]),
            'org_id': r[9], 'org_name': r[10],
            'refund_amount': round(refund, 2)
        })
    stats['expiring_savings'] = expiring_savings
    stats['expiring_savings_total'] = round(total_refund, 2)

    return stats

def hash_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def generate_sms_code():
    return '%06d' % (secrets.randbelow(900000) + 100000)

def generate_token():
    return secrets.token_hex(32)

def get_session_user(headers, cur):
    token = (headers or {}).get('X-Auth-Token') or (headers or {}).get('x-auth-token', '')
    if not token:
        return None
    cur.execute("SELECT cs.user_id, u.member_id, u.name, u.phone, u.role FROM client_sessions cs JOIN users u ON u.id=cs.user_id WHERE cs.token='%s' AND cs.expires_at > NOW()" % esc(token))
    row = cur.fetchone()
    if not row:
        return None
    return {'user_id': row[0], 'member_id': row[1], 'name': row[2], 'phone': row[3], 'role': row[4]}

def get_staff_session(params, headers, cur):
    token = (headers or {}).get('X-Auth-Token') or (headers or {}).get('x-auth-token', '')
    if not token:
        token = params.get('staff_token', '')
    if not token:
        return None
    cur.execute("SELECT cs.user_id, u.name, u.role, u.login FROM client_sessions cs JOIN users u ON u.id=cs.user_id WHERE cs.token='%s' AND cs.expires_at > NOW() AND u.role IN ('admin','manager')" % esc(token))
    row = cur.fetchone()
    if not row:
        return None
    return {'user_id': row[0], 'name': row[1], 'role': row[2], 'login': row[3]}

def handle_staff_auth(body, cur, conn, ip=''):
    action = body.get('action', '')

    if action == 'login':
        login = body.get('login', '').strip()
        password = body.get('password', '')
        if not login or not password:
            return {'error': 'Введите логин и пароль'}
        pw_hash = hash_password(password)
        cur.execute("SELECT id, name, role, login FROM users WHERE login='%s' AND password_hash='%s' AND role IN ('admin','manager') AND status='active'" % (esc(login), pw_hash))
        row = cur.fetchone()
        if not row:
            audit_log(cur, None, 'login_failed', 'auth', None, login, '', ip)
            conn.commit()
            return {'error': 'Неверный логин или пароль'}
        user_id, name, role, ulogin = row
        token = generate_token()
        expires = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
        cur.execute("INSERT INTO client_sessions (user_id, token, expires_at) VALUES (%s,'%s','%s')" % (user_id, token, expires))
        cur.execute("UPDATE users SET last_login=NOW() WHERE id=%s" % user_id)
        staff_info = {'user_id': user_id, 'name': name, 'role': role}
        audit_log(cur, staff_info, 'login', 'auth', user_id, ulogin, '', ip)
        conn.commit()
        return {'success': True, 'token': token, 'user': {'name': name, 'role': role, 'login': ulogin}}

    if action == 'check':
        token = body.get('token', '')
        cur.execute("SELECT u.id, u.name, u.role, u.login FROM users u JOIN client_sessions cs ON cs.user_id=u.id WHERE cs.token='%s' AND cs.expires_at > NOW() AND u.role IN ('admin','manager')" % esc(token))
        row = cur.fetchone()
        if not row:
            return {'_status': 401, 'error': 'Не авторизован'}
        return {'success': True, 'user': {'name': row[1], 'role': row[2], 'login': row[3]}}

    if action == 'logout':
        token = body.get('token', '')
        cur.execute("UPDATE client_sessions SET expires_at=NOW() WHERE token='%s'" % esc(token))
        conn.commit()
        return {'success': True}

    if action == 'change_password':
        token = body.get('token', '')
        old_pw = body.get('old_password', '')
        new_pw = body.get('new_password', '')
        if not new_pw or len(new_pw) < 6:
            return {'error': 'Новый пароль не менее 6 символов'}
        cur.execute("SELECT cs.user_id FROM client_sessions cs WHERE cs.token='%s' AND cs.expires_at > NOW()" % esc(token))
        row = cur.fetchone()
        if not row:
            return {'_status': 401, 'error': 'Сессия истекла'}
        user_id = row[0]
        cur.execute("SELECT password_hash FROM users WHERE id=%s" % user_id)
        cur_hash = cur.fetchone()[0]
        if cur_hash and cur_hash != hash_password(old_pw):
            return {'error': 'Неверный текущий пароль'}
        cur.execute("UPDATE users SET password_hash='%s' WHERE id=%s" % (hash_password(new_pw), user_id))
        conn.commit()
        return {'success': True}

    return {'error': 'Неизвестное действие'}

def handle_users(method, params, body, staff, cur, conn):
    if staff['role'] != 'admin':
        return {'_status': 403, 'error': 'Только администратор может управлять пользователями'}

    if method == 'GET':
        user_id = params.get('id')
        if user_id:
            return query_one(cur, "SELECT id, login, name, email, phone, role, status, member_id, last_login, created_at FROM users WHERE id=%s" % user_id)
        return query_rows(cur, "SELECT id, login, name, email, phone, role, status, member_id, last_login, created_at FROM users ORDER BY created_at DESC")

    elif method == 'POST':
        action = body.get('action', 'create')
        if action == 'create':
            login = body.get('login', '').strip()
            name = body.get('name', '').strip()
            role = body.get('role', 'manager')
            password = body.get('password', '')
            email = body.get('email', '')
            phone = body.get('phone', '')
            member_id = body.get('member_id')
            if not login or not name:
                return {'error': 'Логин и имя обязательны'}
            if role not in ('admin', 'manager', 'client'):
                return {'error': 'Недопустимая роль'}
            if role == 'client' and not member_id:
                return {'error': 'Для клиента необходимо указать пайщика'}
            if not password or len(password) < 6:
                return {'error': 'Пароль не менее 6 символов'}
            cur.execute("SELECT id FROM users WHERE login='%s'" % esc(login))
            if cur.fetchone():
                return {'error': 'Логин уже занят'}
            if member_id:
                cur.execute("SELECT id FROM users WHERE member_id=%s AND role='client'" % int(member_id))
                if cur.fetchone():
                    return {'error': 'У этого пайщика уже есть учётная запись'}
            pw_hash = hash_password(password)
            mid_sql = str(int(member_id)) if member_id else 'NULL'
            cur.execute("INSERT INTO users (login, name, email, phone, role, password_hash, member_id) VALUES ('%s','%s','%s','%s','%s','%s',%s) RETURNING id" % (esc(login), esc(name), esc(email), esc(phone), role, pw_hash, mid_sql))
            uid = cur.fetchone()[0]
            audit_log(cur, staff, 'create', 'user', uid, '%s (%s)' % (login, role), 'member_id: %s' % mid_sql, '')
            conn.commit()
            return {'id': uid, 'login': login}
        elif action == 'update':
            uid = body.get('id')
            if not uid:
                return {'error': 'Укажите id'}
            updates = []
            for f in ['name', 'email', 'phone', 'role', 'status', 'login']:
                if f in body and body[f] is not None:
                    if f == 'role' and body[f] not in ('admin', 'manager', 'client'):
                        return {'error': 'Недопустимая роль'}
                    updates.append("%s='%s'" % (f, esc(body[f])))
            if 'member_id' in body:
                mid = body['member_id']
                if mid:
                    cur.execute("SELECT id FROM users WHERE member_id=%s AND role='client' AND id!=%s" % (int(mid), uid))
                    if cur.fetchone():
                        return {'error': 'У этого пайщика уже есть учётная запись'}
                    updates.append("member_id=%s" % int(mid))
                else:
                    updates.append("member_id=NULL")
            if body.get('password'):
                if len(body['password']) < 6:
                    return {'error': 'Пароль не менее 6 символов'}
                updates.append("password_hash='%s'" % hash_password(body['password']))
            if updates:
                cur.execute("UPDATE users SET %s WHERE id=%s" % (', '.join(updates), uid))
                changed = [f for f in body if f not in ('id', 'entity', 'action', 'password')]
                if body.get('password'):
                    changed.append('password')
                audit_log(cur, staff, 'update', 'user', uid, '', ', '.join(changed), '')
                conn.commit()
            return {'success': True}
        elif action == 'delete':
            uid = body.get('id')
            if not uid:
                return {'error': 'Укажите id'}
            if int(uid) == staff['user_id']:
                return {'error': 'Нельзя удалить самого себя'}
            cur.execute("UPDATE users SET status='blocked' WHERE id=%s" % uid)
            cur.execute("UPDATE client_sessions SET expires_at=NOW() WHERE user_id=%s" % uid)
            audit_log(cur, staff, 'block', 'user', uid, '', '', '')
            conn.commit()
            return {'success': True}
        elif action == 'bulk_create_clients':
            default_password = body.get('password', 'kpk12345')
            if len(default_password) < 6:
                return {'error': 'Пароль не менее 6 символов'}
            pw_hash = hash_password(default_password)
            cur.execute("SELECT m.id, m.member_no, m.phone, CASE WHEN m.member_type='FL' THEN CONCAT(m.last_name,' ',m.first_name) ELSE m.company_name END as name FROM members m WHERE m.status='active' AND NOT EXISTS (SELECT 1 FROM users u WHERE u.member_id=m.id AND u.role='client')")
            rows = cur.fetchall()
            created = 0
            for r in rows:
                mid, mno, mphone, mname = r[0], r[1], r[2] or '', r[3] or 'Клиент'
                login = mno.lower().replace('-', '').replace(' ', '')
                cur.execute("SELECT id FROM users WHERE login='%s'" % esc(login))
                if cur.fetchone():
                    login = login + str(mid)
                cur.execute("INSERT INTO users (login, name, email, phone, role, password_hash, member_id) VALUES ('%s','%s','','%s','client','%s',%s)" % (esc(login), esc(mname), esc(mphone), pw_hash, mid))
                created += 1
            if created > 0:
                audit_log(cur, staff, 'bulk_create_clients', 'user', None, '', 'Создано: %s' % created, '')
                conn.commit()
            return {'success': True, 'created': created, 'password': default_password}
    return {'error': 'Неизвестное действие'}

def handle_auth(method, body, cur, conn):
    action = body.get('action', '')

    if action == 'send_sms':
        phone = body.get('phone', '').strip()
        if not phone:
            return {'error': 'Укажите номер телефона'}
        clean_phone = ''.join(c for c in phone if c.isdigit())
        if len(clean_phone) == 11 and clean_phone[0] == '8':
            clean_phone = '7' + clean_phone[1:]

        cur.execute("SELECT m.id, m.phone FROM members m WHERE REPLACE(REPLACE(REPLACE(REPLACE(m.phone,' ',''),'-',''),'(',''),')','') LIKE '%%%s%%' AND m.status='active'" % clean_phone[-10:])
        member = cur.fetchone()
        if not member:
            return {'error': 'Пайщик с таким номером не найден. Обратитесь в КПК.'}

        member_id = member[0]
        cur.execute("SELECT id, password_hash FROM users WHERE member_id=%s AND role='client'" % member_id)
        user_row = cur.fetchone()

        code = generate_sms_code()
        expires = (datetime.now() + timedelta(minutes=5)).strftime('%Y-%m-%d %H:%M:%S')

        if user_row:
            user_id = user_row[0]
            has_password = bool(user_row[1])
            cur.execute("UPDATE users SET sms_code='%s', sms_code_expires='%s' WHERE id=%s" % (code, expires, user_id))
        else:
            cur.execute("SELECT CASE WHEN member_type='FL' THEN CONCAT(last_name,' ',first_name) ELSE company_name END FROM members WHERE id=%s" % member_id)
            name_row = cur.fetchone()
            uname = name_row[0] if name_row else 'Клиент'
            cur.execute("INSERT INTO users (member_id, name, email, phone, role, sms_code, sms_code_expires) VALUES (%s,'%s','','%s','client','%s','%s') RETURNING id" % (member_id, esc(uname), esc(phone), code, expires))
            user_id = cur.fetchone()[0]
            has_password = False

        conn.commit()
        return {'success': True, 'has_password': has_password, 'sms_sent': True, 'debug_code': code}

    elif action == 'verify_sms':
        phone = body.get('phone', '').strip()
        code = body.get('code', '').strip()
        clean_phone = ''.join(c for c in phone if c.isdigit())

        cur.execute("SELECT u.id, u.password_hash, u.name, u.member_id FROM users u JOIN members m ON m.id=u.member_id WHERE REPLACE(REPLACE(REPLACE(REPLACE(m.phone,' ',''),'-',''),'(',''),')','') LIKE '%%%s%%' AND u.role='client' AND u.sms_code='%s' AND u.sms_code_expires > NOW()" % (clean_phone[-10:], esc(code)))
        row = cur.fetchone()
        if not row:
            return {'error': 'Неверный код или код истёк'}

        user_id, pw_hash, name, member_id = row[0], row[1], row[2], row[3]
        has_password = bool(pw_hash)

        cur.execute("UPDATE users SET sms_code=NULL, sms_code_expires=NULL WHERE id=%s" % user_id)

        if has_password:
            token = generate_token()
            expires = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
            cur.execute("INSERT INTO client_sessions (user_id, token, expires_at) VALUES (%s,'%s','%s')" % (user_id, token, expires))
            cur.execute("UPDATE users SET last_login=NOW() WHERE id=%s" % user_id)
            conn.commit()
            return {'success': True, 'has_password': True, 'authenticated': True, 'token': token, 'user': {'name': name, 'member_id': member_id}}
        else:
            temp_token = generate_token()
            expires = (datetime.now() + timedelta(minutes=15)).strftime('%Y-%m-%d %H:%M:%S')
            cur.execute("INSERT INTO client_sessions (user_id, token, expires_at) VALUES (%s,'%s','%s')" % (user_id, temp_token, expires))
            conn.commit()
            return {'success': True, 'has_password': False, 'setup_token': temp_token}

    elif action == 'set_password':
        token = body.get('setup_token') or body.get('token', '')
        password = body.get('password', '')
        if not password or len(password) < 6:
            return {'error': 'Пароль должен быть не менее 6 символов'}

        cur.execute("SELECT cs.user_id FROM client_sessions cs WHERE cs.token='%s' AND cs.expires_at > NOW()" % esc(token))
        row = cur.fetchone()
        if not row:
            return {'error': 'Сессия истекла, повторите авторизацию'}

        user_id = row[0]
        pw_hash = hash_password(password)
        cur.execute("UPDATE users SET password_hash='%s' WHERE id=%s" % (pw_hash, user_id))

        cur.execute("UPDATE client_sessions SET expires_at=NOW() WHERE token='%s'" % esc(token))

        new_token = generate_token()
        expires = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
        cur.execute("INSERT INTO client_sessions (user_id, token, expires_at) VALUES (%s,'%s','%s')" % (user_id, new_token, expires))
        cur.execute("UPDATE users SET last_login=NOW() WHERE id=%s" % user_id)

        cur.execute("SELECT name, member_id FROM users WHERE id=%s" % user_id)
        ur = cur.fetchone()
        conn.commit()
        return {'success': True, 'token': new_token, 'user': {'name': ur[0], 'member_id': ur[1]}}

    elif action == 'login_password':
        phone = body.get('phone', '').strip()
        login = body.get('login', '').strip()
        password = body.get('password', '')
        pw_hash = hash_password(password)

        row = None
        if login:
            cur.execute("SELECT u.id, u.name, u.member_id FROM users u WHERE u.login='%s' AND u.role='client' AND u.password_hash='%s' AND u.status='active'" % (esc(login), pw_hash))
            row = cur.fetchone()
        if not row and phone:
            clean_phone = ''.join(c for c in phone if c.isdigit())
            cur.execute("SELECT u.id, u.name, u.member_id FROM users u JOIN members m ON m.id=u.member_id WHERE REPLACE(REPLACE(REPLACE(REPLACE(m.phone,' ',''),'-',''),'(',''),')','') LIKE '%%%s%%' AND u.role='client' AND u.password_hash='%s'" % (clean_phone[-10:], pw_hash))
            row = cur.fetchone()
        if not row:
            return {'error': 'Неверный логин/телефон или пароль'}

        user_id, name, member_id = row
        token = generate_token()
        expires = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
        cur.execute("INSERT INTO client_sessions (user_id, token, expires_at) VALUES (%s,'%s','%s')" % (user_id, token, expires))
        cur.execute("UPDATE users SET last_login=NOW() WHERE id=%s" % user_id)
        conn.commit()
        return {'success': True, 'token': token, 'user': {'name': name, 'member_id': member_id}}

    elif action == 'change_password':
        token = body.get('token', '')
        old_pw = body.get('old_password', '')
        new_pw = body.get('new_password', '')
        if not new_pw or len(new_pw) < 6:
            return {'error': 'Новый пароль должен быть не менее 6 символов'}

        cur.execute("SELECT cs.user_id FROM client_sessions cs WHERE cs.token='%s' AND cs.expires_at > NOW()" % esc(token))
        row = cur.fetchone()
        if not row:
            return {'error': 'Сессия истекла'}
        user_id = row[0]

        cur.execute("SELECT password_hash FROM users WHERE id=%s" % user_id)
        cur_hash = cur.fetchone()[0]
        if cur_hash and cur_hash != hash_password(old_pw):
            return {'error': 'Неверный текущий пароль'}

        cur.execute("UPDATE users SET password_hash='%s' WHERE id=%s" % (hash_password(new_pw), user_id))
        conn.commit()
        return {'success': True}

    elif action == 'logout':
        token = body.get('token', '')
        cur.execute("UPDATE client_sessions SET expires_at=NOW() WHERE token='%s'" % esc(token))
        conn.commit()
        return {'success': True}

    elif action == 'check':
        token = body.get('token', '')
        cur.execute("SELECT u.id, u.name, u.member_id FROM users u JOIN client_sessions cs ON cs.user_id=u.id WHERE cs.token='%s' AND cs.expires_at > NOW()" % esc(token))
        row = cur.fetchone()
        if not row:
            return {'error': 'Не авторизован'}
        return {'success': True, 'user': {'name': row[1], 'member_id': row[2]}}

    return {'error': 'Неизвестное действие'}

def handle_cabinet(method, params, body, headers, cur):
    token = params.get('token') or body.get('token', '')
    if not token:
        token = (headers or {}).get('X-Auth-Token') or (headers or {}).get('x-auth-token', '')

    cur.execute("SELECT u.id, u.member_id FROM users u JOIN client_sessions cs ON cs.user_id=u.id WHERE cs.token='%s' AND cs.expires_at > NOW()" % esc(token))
    row = cur.fetchone()
    if not row:
        return {'_status': 401, 'error': 'Не авторизован'}
    member_id = row[1]

    action = params.get('action') or body.get('action', 'overview')

    if action == 'overview':
        cur.execute("SELECT CASE WHEN member_type='FL' THEN CONCAT(last_name,' ',first_name,' ',middle_name) ELSE company_name END as name, member_no, phone, email FROM members WHERE id=%s" % member_id)
        mr = cur.fetchone()
        info = {'name': mr[0], 'member_no': mr[1], 'phone': mr[2], 'email': mr[3]} if mr else {}

        loans = query_rows(cur, """
            SELECT l.id, l.contract_no, l.amount, l.rate, l.term_months, l.schedule_type, l.start_date, l.end_date,
                   l.monthly_payment, l.balance, l.status, l.org_id,
                   o.name as org_name, o.short_name as org_short_name
            FROM loans l LEFT JOIN organizations o ON o.id=l.org_id
            WHERE l.member_id=%s ORDER BY l.created_at DESC
        """ % member_id)

        savings = query_rows(cur, """
            SELECT s.id, s.contract_no, s.amount, s.rate, s.term_months, s.payout_type, s.start_date, s.end_date,
                   s.accrued_interest, s.paid_interest, s.current_balance, s.status, s.org_id,
                   o.name as org_name, o.short_name as org_short_name
            FROM savings s LEFT JOIN organizations o ON o.id=s.org_id
            WHERE s.member_id=%s ORDER BY s.created_at DESC
        """ % member_id)

        shares = query_rows(cur, """
            SELECT sa.id, sa.account_no, sa.balance, sa.total_in, sa.total_out, sa.status, sa.org_id,
                   o.name as org_name, o.short_name as org_short_name
            FROM share_accounts sa LEFT JOIN organizations o ON o.id=sa.org_id
            WHERE sa.member_id=%s ORDER BY sa.created_at DESC
        """ % member_id)

        org_ids = set()
        for item in loans + savings + shares:
            if item.get('org_id'):
                org_ids.add(item['org_id'])
        orgs_map = {}
        if org_ids:
            org_rows = query_rows(cur, "SELECT id, name, short_name, inn, kpp, bank_name, bik, rs, ks FROM organizations WHERE id IN (%s)" % ','.join(str(i) for i in org_ids))
            for o in org_rows:
                orgs_map[o['id']] = o

        return {'info': info, 'loans': loans, 'savings': savings, 'shares': shares, 'organizations': orgs_map}

    elif action == 'loan_detail':
        loan_id = params.get('id') or body.get('id')
        loan = query_one(cur, "SELECT * FROM loans WHERE id=%s AND member_id=%s" % (loan_id, member_id))
        if not loan:
            return {'error': 'Договор не найден'}
        loan['schedule'] = query_rows(cur, "SELECT * FROM loan_schedule WHERE loan_id=%s ORDER BY payment_no" % loan_id)
        loan['payments'] = query_rows(cur, "SELECT * FROM loan_payments WHERE loan_id=%s ORDER BY payment_date" % loan_id)
        return loan

    elif action == 'saving_detail':
        saving_id = params.get('id') or body.get('id')
        saving = query_one(cur, "SELECT * FROM savings WHERE id=%s AND member_id=%s" % (saving_id, member_id))
        if not saving:
            return {'error': 'Договор не найден'}
        saving['schedule'] = query_rows(cur, "SELECT * FROM savings_schedule WHERE saving_id=%s ORDER BY period_no" % saving_id)
        cur.execute("SELECT COALESCE(SUM(daily_amount), 0) FROM savings_daily_accruals WHERE saving_id=%s" % saving_id)
        saving['total_daily_accrued'] = float(cur.fetchone()[0])
        return saving

    return {'error': 'Неизвестное действие'}

def handle_audit(params, staff, cur):
    if staff.get('role') != 'admin':
        return {'_status': 403, 'error': 'Только администратор может просматривать журнал'}
    limit = min(int(params.get('limit', 100)), 500)
    offset = int(params.get('offset', 0))
    entity_filter = params.get('filter_entity', '')
    action_filter = params.get('filter_action', '')
    where = []
    if entity_filter:
        where.append("entity='%s'" % esc(entity_filter))
    if action_filter:
        where.append("action='%s'" % esc(action_filter))
    where_sql = (' WHERE ' + ' AND '.join(where)) if where else ''
    cur.execute("SELECT COUNT(*) FROM audit_log%s" % where_sql)
    total = cur.fetchone()[0]
    rows = query_rows(cur, "SELECT * FROM audit_log%s ORDER BY created_at DESC LIMIT %s OFFSET %s" % (where_sql, limit, offset))
    return {'items': rows, 'total': total}

def handle_org_settings(method, body, staff, cur, conn):
    if staff['role'] != 'admin':
        return {'_status': 403, 'error': 'Только администратор может управлять настройками'}
    if method == 'GET':
        cur.execute("SELECT key, value FROM organization_settings ORDER BY id")
        rows = cur.fetchall()
        return {r[0]: r[1] for r in rows}
    elif method == 'POST':
        data = body.get('settings', {})
        allowed = ('name', 'inn', 'ogrn', 'director_fio', 'bank_name', 'bik', 'rs', 'phone', 'website', 'email', 'telegram', 'whatsapp')
        for k, v in data.items():
            if k in allowed:
                cur.execute("INSERT INTO organization_settings (key, value, updated_at) VALUES ('%s', '%s', NOW()) ON CONFLICT (key) DO UPDATE SET value='%s', updated_at=NOW()" % (esc(k), esc(v), esc(v)))
        audit_log(cur, staff, 'update', 'org_settings', None, '', ', '.join(data.keys()), '')
        conn.commit()
        return {'success': True}
    return {'error': 'Неизвестный метод'}

def handle_organizations(method, params, body, staff, cur, conn, ip=''):
    if method == 'GET':
        org_id = params.get('id')
        if org_id:
            return query_one(cur, "SELECT * FROM organizations WHERE id=%s AND is_active=true" % org_id)
        return query_rows(cur, "SELECT * FROM organizations WHERE is_active=true ORDER BY name")

    elif method == 'POST':
        action = body.get('action', 'create')

        if action == 'create':
            if staff.get('role') != 'admin':
                return {'_status': 403, 'error': 'Только администратор может создавать организации'}
            cur.execute("""
                INSERT INTO organizations (name, short_name, inn, ogrn, kpp, director_fio, director_position,
                    legal_address, actual_address, bank_name, bik, rs, ks, phone, email, website, telegram, whatsapp, logo_url)
                VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')
                RETURNING id
            """ % (
                esc(body.get('name', '')), esc(body.get('short_name', '')), esc(body.get('inn', '')),
                esc(body.get('ogrn', '')), esc(body.get('kpp', '')), esc(body.get('director_fio', '')),
                esc(body.get('director_position', '')), esc(body.get('legal_address', '')),
                esc(body.get('actual_address', '')), esc(body.get('bank_name', '')),
                esc(body.get('bik', '')), esc(body.get('rs', '')), esc(body.get('ks', '')),
                esc(body.get('phone', '')), esc(body.get('email', '')), esc(body.get('website', '')),
                esc(body.get('telegram', '')), esc(body.get('whatsapp', '')), esc(body.get('logo_url', ''))
            ))
            org_id = cur.fetchone()[0]
            audit_log(cur, staff, 'create', 'organization', org_id, esc(body.get('name', '')), '', ip)
            conn.commit()
            return {'id': org_id}

        elif action == 'update':
            if staff.get('role') != 'admin':
                return {'_status': 403, 'error': 'Только администратор может редактировать организации'}
            org_id = body.get('id')
            if not org_id:
                return {'error': 'Не указан id организации'}
            fields = []
            allowed = ('name', 'short_name', 'inn', 'ogrn', 'kpp', 'director_fio', 'director_position',
                        'legal_address', 'actual_address', 'bank_name', 'bik', 'rs', 'ks',
                        'phone', 'email', 'website', 'telegram', 'whatsapp', 'logo_url')
            for k in allowed:
                if k in body:
                    fields.append("%s='%s'" % (k, esc(body[k])))
            if fields:
                fields.append("updated_at=NOW()")
                cur.execute("UPDATE organizations SET %s WHERE id=%s" % (', '.join(fields), org_id))
            audit_log(cur, staff, 'update', 'organization', org_id, esc(body.get('name', '')), '', ip)
            conn.commit()
            return {'success': True}

        elif action == 'upload_logo':
            if staff.get('role') != 'admin':
                return {'_status': 403, 'error': 'Только администратор может загружать логотип'}
            org_id = body.get('id')
            if not org_id:
                return {'error': 'Не указан id организации'}
            logo_b64 = body.get('logo')
            if not logo_b64:
                return {'error': 'Не передан логотип'}
            allowed_types = {'image/png': 'png', 'image/jpeg': 'jpg', 'image/webp': 'webp', 'image/svg+xml': 'svg'}
            ct = body.get('content_type', 'image/png')
            if ct not in allowed_types:
                return {'_status': 400, 'error': 'Недопустимый формат. Допустимы: PNG, JPEG, WebP, SVG'}
            import boto3
            logo_data = base64.b64decode(logo_b64)
            if len(logo_data) > 2 * 1024 * 1024:
                return {'_status': 400, 'error': 'Файл слишком большой. Максимум 2 МБ'}
            ext = allowed_types[ct]
            s3 = boto3.client('s3',
                endpoint_url='https://bucket.poehali.dev',
                aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
                aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'])
            s3_key = 'org_logos/%s.%s' % (org_id, ext)
            s3.put_object(Bucket='files', Key=s3_key, Body=logo_data, ContentType=ct)
            cdn_url = 'https://cdn.poehali.dev/projects/%s/bucket/%s' % (os.environ['AWS_ACCESS_KEY_ID'], s3_key)
            cur.execute("UPDATE organizations SET logo_url='%s', updated_at=NOW() WHERE id=%s" % (esc(cdn_url), org_id))
            audit_log(cur, staff, 'upload_logo', 'organization', org_id, '', '', ip)
            conn.commit()
            return {'success': True, 'logo_url': cdn_url}

        elif action == 'delete':
            if staff.get('role') != 'admin':
                return {'_status': 403, 'error': 'Только администратор может удалять организации'}
            org_id = body.get('id')
            if not org_id:
                return {'error': 'Не указан id организации'}
            cur.execute("UPDATE organizations SET is_active=false, updated_at=NOW() WHERE id=%s" % org_id)
            audit_log(cur, staff, 'delete', 'organization', org_id, '', '', ip)
            conn.commit()
            return {'success': True}

    return {'error': 'Неизвестный метод'}

PROTECTED_ENTITIES = {'dashboard', 'members', 'loans', 'savings', 'shares', 'export', 'users', 'audit', 'org_settings', 'organizations'}

def handler(event, context):
    """Единый API для ERP кредитного кооператива: пайщики, займы, сбережения, паевые счета, ЛК, авторизация"""
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': {'Access-Control-Allow-Origin': '*', 'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS', 'Access-Control-Allow-Headers': 'Content-Type, X-Auth-Token', 'Access-Control-Max-Age': '86400'}, 'body': ''}

    headers = {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'}
    method = event.get('httpMethod', 'GET')
    params = event.get('queryStringParameters') or {}
    body = json.loads(event.get('body', '{}')) if event.get('body') else {}
    ev_headers = event.get('headers') or {}

    entity = params.get('entity') or body.get('entity', 'dashboard')

    src_ip = (event.get('requestContext') or {}).get('identity', {}).get('sourceIp', '')

    conn = get_conn()
    cur = conn.cursor()

    try:
        staff = None
        if entity in PROTECTED_ENTITIES:
            staff = get_staff_session(params, ev_headers, cur)
            if not staff:
                return {'statusCode': 401, 'headers': headers, 'body': json.dumps({'error': 'Требуется авторизация'})}

            if staff['role'] == 'manager':
                if method == 'DELETE':
                    return {'statusCode': 403, 'headers': headers, 'body': json.dumps({'error': 'Менеджер не может удалять записи'})}
                if entity in ('users', 'audit', 'org_settings'):
                    return {'statusCode': 403, 'headers': headers, 'body': json.dumps({'error': 'Недостаточно прав'})}
                action = params.get('action') or body.get('action', '')
                if action and 'delete' in action:
                    return {'statusCode': 403, 'headers': headers, 'body': json.dumps({'error': 'Менеджер не может удалять записи'})}

        if entity == 'dashboard':
            result = handle_dashboard(cur, params)
        elif entity == 'members':
            result = handle_members(method, params, body, cur, conn, staff, src_ip)
        elif entity == 'loans':
            result = handle_loans(method, params, body, cur, conn, staff, src_ip)
        elif entity == 'savings':
            result = handle_savings(method, params, body, cur, conn, staff, src_ip)
        elif entity == 'shares':
            result = handle_shares(method, params, body, cur, conn, staff, src_ip)
        elif entity == 'export':
            result = handle_export(params, cur)
        elif entity == 'users':
            result = handle_users(method, params, body, staff, cur, conn)
        elif entity == 'audit':
            result = handle_audit(params, staff, cur)
        elif entity == 'org_settings':
            result = handle_org_settings(method, body, staff, cur, conn)
        elif entity == 'organizations':
            result = handle_organizations(method, params, body, staff, cur, conn, src_ip)
        elif entity == 'staff_auth':
            result = handle_staff_auth(body, cur, conn, src_ip)
        elif entity == 'auth':
            result = handle_auth(method, body, cur, conn)
        elif entity == 'cabinet':
            result = handle_cabinet(method, params, body, ev_headers, cur)
        elif entity == 'cron':
            cron_action = body.get('action') or params.get('action', '')
            if cron_action == 'daily_accrue':
                result = handle_savings('POST', params, {'action': 'daily_accrue', 'date': body.get('date', date.today().isoformat())}, cur, conn, None, src_ip)
            else:
                result = {'error': 'Неизвестное cron действие'}
        else:
            return {'statusCode': 400, 'headers': headers, 'body': json.dumps({'error': 'Unknown entity: %s' % entity})}

        if isinstance(result, dict) and '_status' in result:
            st = result.pop('_status')
            return {'statusCode': st, 'headers': headers, 'body': json.dumps(result)}

        if result is None:
            return {'statusCode': 404, 'headers': headers, 'body': json.dumps({'error': 'Не найдено'})}

        if isinstance(result, dict) and 'error' in result:
            return {'statusCode': 400, 'headers': headers, 'body': json.dumps(result)}

        code = 201 if method == 'POST' else 200
        return {'statusCode': code, 'headers': headers, 'body': json.dumps(result)}
    except Exception as e:
        conn.rollback()
        err = str(e)
        msg = err
        if 'unique' in err.lower() and 'constraint' in err.lower():
            msg = 'Запись с такими данными уже существует'
        elif 'check' in err.lower() and 'constraint' in err.lower():
            msg = 'Некорректные данные. Проверьте заполнение полей'
        elif 'foreign key' in err.lower() or 'not-null' in err.lower():
            msg = 'Невозможно выполнить: связанные данные не найдены или не заполнены обязательные поля'
        elif 'numeric' in err.lower() and 'overflow' in err.lower():
            msg = 'Слишком большое число. Проверьте суммы и ставки'
        elif 'division by zero' in err.lower():
            msg = 'Ошибка вычисления: деление на ноль. Проверьте параметры'
        return {'statusCode': 500, 'headers': headers, 'body': json.dumps({'error': msg})}
    finally:
        cur.close()
        conn.close()